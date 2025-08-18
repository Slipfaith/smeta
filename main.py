import sys
import os
from datetime import datetime
from typing import Dict, List, Any
import json

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QGroupBox, QComboBox, QSpinBox,
    QDoubleSpinBox, QTableWidget, QTableWidgetItem, QTabWidget,
    QScrollArea, QMessageBox, QFileDialog, QTextEdit, QCheckBox,
    QHeaderView, QFrame, QSplitter
)
from PySide6.QtCore import Qt, QThread, QObject, Signal
from PySide6.QtGui import QFont, QIcon, QPalette

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Необходимо установить openpyxl: pip install openpyxl")
    sys.exit(1)


class ServiceConfig:
    """Конфигурация услуг и их параметров"""

    TRANSLATION_ROWS = [
        {"name": "Перевод, новые слова (100%)", "multiplier": 1.0, "is_base": True},
        {"name": "Перевод, совпадения 75-94% (66%)", "multiplier": 0.66, "is_base": False},
        {"name": "Перевод, совпадения 95-99% (33%)", "multiplier": 0.33, "is_base": False},
        {"name": "Перевод, повторы и 100% совпадения (30%)", "multiplier": 0.30, "is_base": False}
    ]

    EDITING_ROWS = [
        {"name": "Редактирование новых слов", "multiplier": 1.0, "is_base": True},
        {"name": "Редактирование повторов", "multiplier": 0.5, "is_base": False},
        {"name": "Корректорская правка", "multiplier": 1.0, "is_base": True},
        {"name": "Стилистическая правка", "multiplier": 1.0, "is_base": True}
    ]

    ADDITIONAL_SERVICES = {
        "Верстка": [
            {"name": "InDesign верстка", "multiplier": 1.0, "is_base": True},
            {"name": "PowerPoint верстка", "multiplier": 1.0, "is_base": True},
            {"name": "PDF верстка", "multiplier": 1.0, "is_base": True},
            {"name": "Графика/Изображения", "multiplier": 1.0, "is_base": True}
        ],
        "Локализация мультимедиа": [
            {"name": "Создание субтитров", "multiplier": 1.0, "is_base": True},
            {"name": "Озвучка", "multiplier": 1.0, "is_base": True},
            {"name": "Видеомонтаж", "multiplier": 1.0, "is_base": True},
            {"name": "Синхронизация", "multiplier": 1.0, "is_base": True}
        ],
        "Тестирование/QA": [
            {"name": "Лингвистическое тестирование", "multiplier": 1.0, "is_base": True},
            {"name": "Функциональное тестирование", "multiplier": 1.0, "is_base": True},
            {"name": "Косметическое тестирование", "multiplier": 1.0, "is_base": True},
            {"name": "Финальная проверка", "multiplier": 1.0, "is_base": True}
        ],
        "Прочие услуги": [
            {"name": "Создание терминологии", "multiplier": 1.0, "is_base": True},
            {"name": "Подготовка Translation Memory", "multiplier": 1.0, "is_base": True},
            {"name": "Анализ CAT-инструмента", "multiplier": 1.0, "is_base": True},
            {"name": "Консультации", "multiplier": 1.0, "is_base": True}
        ]
    }


class LanguagePairWidget(QWidget):
    """Виджет для одной языковой пары"""

    def __init__(self, pair_name: str):
        super().__init__()
        self.pair_name = pair_name
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()

        # Заголовок языковой пары
        title = QLabel(f"Языковая пара: {self.pair_name}")
        title.setFont(QFont("Arial", 10, QFont.Bold))
        layout.addWidget(title)

        # Услуги для этой языковой пары
        self.services_layout = QVBoxLayout()

        # Перевод
        self.translation_group = self.create_service_group("Перевод", ServiceConfig.TRANSLATION_ROWS)
        self.services_layout.addWidget(self.translation_group)

        # Редактирование
        self.editing_group = self.create_service_group("Редактирование", ServiceConfig.EDITING_ROWS)
        self.services_layout.addWidget(self.editing_group)

        layout.addLayout(self.services_layout)
        self.setLayout(layout)

    def create_service_group(self, service_name: str, rows: List[Dict]) -> QGroupBox:
        """Создает группу для услуги с таблицей параметров"""
        group = QGroupBox(service_name)
        group.setCheckable(True)
        group.setChecked(False)

        layout = QVBoxLayout()

        # Таблица параметров
        table = QTableWidget(len(rows), 4)  # строки, колонки: Параметр, Объем, Ставка, Сумма
        table.setHorizontalHeaderLabels(["Параметр", "Объем", "Ставка (руб)", "Сумма (руб)"])

        # Сохраняем базовую ставку для автоматических расчетов
        base_rate_row = None

        for i, row_info in enumerate(rows):
            # Название параметра
            table.setItem(i, 0, QTableWidgetItem(row_info["name"]))

            # Объем
            volume_item = QTableWidgetItem("0")
            table.setItem(i, 1, volume_item)

            # Ставка
            rate_item = QTableWidgetItem("0.00")
            if not row_info["is_base"]:
                rate_item.setFlags(Qt.ItemIsEnabled)  # только чтение для автоматических ставок
            else:
                if base_rate_row is None:
                    base_rate_row = i  # запоминаем первую базовую ставку
            table.setItem(i, 2, rate_item)

            # Сумма (только чтение)
            sum_item = QTableWidgetItem("0.00")
            sum_item.setFlags(Qt.ItemIsEnabled)  # только чтение
            table.setItem(i, 3, sum_item)

        # Подключаем обновление ставок и сумм при изменении данных
        table.itemChanged.connect(lambda item: self.update_rates_and_sums(table, rows, base_rate_row))

        # Настройка ширины колонок
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)

        layout.addWidget(table)
        group.setLayout(layout)

        # Сохраняем ссылки на таблицу и конфигурацию строк
        setattr(group, 'table', table)
        setattr(group, 'rows_config', rows)
        setattr(group, 'base_rate_row', base_rate_row)

        return group

    def update_rates_and_sums(self, table: QTableWidget, rows: List[Dict], base_rate_row: int):
        """Обновляет ставки и суммы в таблице"""
        try:
            # Получаем базовую ставку
            base_rate = 0.0
            if base_rate_row is not None:
                base_rate = float(table.item(base_rate_row, 2).text() or "0")

            # Обновляем все строки
            for row in range(table.rowCount()):
                row_config = rows[row]

                # Обновляем ставку для неосновных строк
                if not row_config["is_base"] and base_rate_row is not None:
                    auto_rate = base_rate * row_config["multiplier"]
                    table.item(row, 2).setText(f"{auto_rate:.2f}")

                # Обновляем сумму
                volume = float(table.item(row, 1).text() or "0")
                rate = float(table.item(row, 2).text() or "0")
                total = volume * rate
                table.item(row, 3).setText(f"{total:.2f}")

        except (ValueError, AttributeError):
            # В случае ошибки просто пропускаем обновление
            pass

    def get_data(self) -> Dict[str, Any]:
        """Получает данные языковой пары"""
        data = {"pair_name": self.pair_name, "services": {}}

        # Перевод
        if self.translation_group.isChecked():
            data["services"]["translation"] = self.get_table_data(self.translation_group.table)

        # Редактирование  
        if self.editing_group.isChecked():
            data["services"]["editing"] = self.get_table_data(self.editing_group.table)

        return data

    def get_table_data(self, table: QTableWidget) -> List[Dict[str, Any]]:
        """Получает данные из таблицы"""
        data = []
        for row in range(table.rowCount()):
            row_data = {
                "parameter": table.item(row, 0).text() if table.item(row, 0) else "",
                "volume": float(table.item(row, 1).text() or "0") if table.item(row, 1) else 0,
                "rate": float(table.item(row, 2).text() or "0") if table.item(row, 2) else 0,
                "total": float(table.item(row, 3).text() or "0") if table.item(row, 3) else 0
            }
            data.append(row_data)
        return data


class AdditionalServicesWidget(QWidget):
    """Виджет для дополнительных услуг"""

    def __init__(self):
        super().__init__()
        self.service_groups = {}
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()

        title = QLabel("Дополнительные услуги")
        title.setFont(QFont("Arial", 12, QFont.Bold))
        layout.addWidget(title)

        # Создаем группы для каждой дополнительной услуги
        for service_name, rows in ServiceConfig.ADDITIONAL_SERVICES.items():
            group = self.create_service_group(service_name, rows)
            self.service_groups[service_name] = group
            layout.addWidget(group)

        self.setLayout(layout)

    def create_service_group(self, service_name: str, rows: List[Dict]) -> QGroupBox:
        """Создает группу для дополнительной услуги"""
        group = QGroupBox(service_name)
        group.setCheckable(True)
        group.setChecked(False)

        layout = QVBoxLayout()

        table = QTableWidget(len(rows), 4)
        table.setHorizontalHeaderLabels(["Параметр", "Объем", "Ставка (руб)", "Сумма (руб)"])

        base_rate_row = None

        for i, row_info in enumerate(rows):
            table.setItem(i, 0, QTableWidgetItem(row_info["name"]))
            table.setItem(i, 1, QTableWidgetItem("0"))

            rate_item = QTableWidgetItem("0.00")
            if not row_info["is_base"]:
                rate_item.setFlags(Qt.ItemIsEnabled)
            else:
                if base_rate_row is None:
                    base_rate_row = i
            table.setItem(i, 2, rate_item)

            sum_item = QTableWidgetItem("0.00")
            sum_item.setFlags(Qt.ItemIsEnabled)
            table.setItem(i, 3, sum_item)

        table.itemChanged.connect(lambda item: self.update_rates_and_sums(table, rows, base_rate_row))

        # Настройка ширины колонок
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)

        layout.addWidget(table)
        group.setLayout(layout)

        setattr(group, 'table', table)
        setattr(group, 'rows_config', rows)
        setattr(group, 'base_rate_row', base_rate_row)

        return group

    def update_rates_and_sums(self, table: QTableWidget, rows: List[Dict], base_rate_row: int):
        """Обновляет ставки и суммы в таблице"""
        try:
            # Получаем базовую ставку
            base_rate = 0.0
            if base_rate_row is not None:
                base_rate = float(table.item(base_rate_row, 2).text() or "0")

            # Обновляем все строки
            for row in range(table.rowCount()):
                row_config = rows[row]

                # Обновляем ставку для неосновных строк
                if not row_config["is_base"] and base_rate_row is not None:
                    auto_rate = base_rate * row_config["multiplier"]
                    table.item(row, 2).setText(f"{auto_rate:.2f}")

                # Обновляем сумму
                volume = float(table.item(row, 1).text() or "0")
                rate = float(table.item(row, 2).text() or "0")
                total = volume * rate
                table.item(row, 3).setText(f"{total:.2f}")

        except (ValueError, AttributeError):
            pass

    def get_data(self) -> Dict[str, Any]:
        """Получает данные дополнительных услуг"""
        data = {}
        for service_name, group in self.service_groups.items():
            if group.isChecked():
                data[service_name] = self.get_table_data(group.table)
        return data

    def get_table_data(self, table: QTableWidget) -> List[Dict[str, Any]]:
        """Получает данные из таблицы"""
        data = []
        for row in range(table.rowCount()):
            row_data = {
                "parameter": table.item(row, 0).text(),
                "volume": float(table.item(row, 1).text() or "0"),
                "rate": float(table.item(row, 2).text() or "0"),
                "total": float(table.item(row, 3).text() or "0")
            }
            data.append(row_data)
        return data


class ExcelExporter:
    """Класс для экспорта данных в Excel"""

    def __init__(self, template_path: str = None):
        self.template_path = template_path

    def export_to_excel(self, project_data: Dict[str, Any], output_path: str):
        """Экспортирует данные проекта в Excel файл на основе шаблона"""
        try:
            # Обязательно загружаем шаблон
            if not self.template_path or not os.path.exists(self.template_path):
                raise Exception("Не указан путь к шаблону Excel или файл не существует")

            # Загружаем шаблон с сохранением формул и форматирования
            wb = load_workbook(self.template_path)
            ws = wb.active

            # Заполняем основные данные проекта в соответствующие ячейки
            self.fill_template_data(ws, project_data)

            # Сохраняем файл
            wb.save(output_path)
            return True

        except Exception as e:
            print(f"Ошибка при экспорте в Excel: {e}")
            return False

    def fill_template_data(self, ws, project_data: Dict[str, Any]):
        """Заполняет данные в шаблон Excel"""

        # Основная информация проекта
        # Ищем и заполняем ячейки с плейсхолдерами
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value)

                    # Заменяем плейсхолдеры на реальные данные
                    if "{{project_name}}" in cell_value:
                        cell.value = cell_value.replace("{{project_name}}", project_data.get('project_name', ''))
                    elif "{{target_langs}}" in cell_value:
                        # Формируем список языковых пар
                        pairs = [pair['pair_name'] for pair in project_data.get('language_pairs', [])]
                        cell.value = cell_value.replace("{{target_langs}}", ", ".join(pairs))
                    elif "{{client}}" in cell_value:
                        cell.value = cell_value.replace("{{client}}", project_data.get('client_name', ''))
                    elif "{{Entity}}" in cell_value:
                        cell.value = cell_value.replace("{{Entity}}", project_data.get('client_name', ''))
                    elif "{{Entity_address}}" in cell_value:
                        cell.value = cell_value.replace("{{Entity_address}}", project_data.get('email', ''))
                    elif "{{client_name}}" in cell_value:
                        cell.value = cell_value.replace("{{client_name}}", project_data.get('contact_person', ''))
                    elif "{{PM_name}}" in cell_value:
                        cell.value = cell_value.replace("{{PM_name}}", "Project Manager")  # или из настроек
                    elif "{{PM_email}}" in cell_value:
                        cell.value = cell_value.replace("{{PM_email}}", "pm@company.com")  # или из настроек

        # Заполняем данные языковых пар
        self.fill_language_pairs_data(ws, project_data)

        # Заполняем дополнительные услуги
        self.fill_additional_services_data(ws, project_data)

    def fill_language_pairs_data(self, ws, project_data: Dict[str, Any]):
        """Заполняет данные языковых пар в существующие таблицы шаблона"""

        language_pairs = project_data.get('language_pairs', [])
        if not language_pairs:
            return

        # Находим все таблицы в шаблоне
        table_sections = self.find_table_sections(ws)

        # Если языковых пар больше, чем таблиц в шаблоне, дублируем последнюю таблицу
        if len(language_pairs) > len(table_sections):
            self.duplicate_table_sections(ws, table_sections, len(language_pairs))
            table_sections = self.find_table_sections(ws)  # Обновляем список после дублирования

        # Заполняем каждую таблицу данными соответствующей языковой пары
        for i, pair_data in enumerate(language_pairs):
            if i < len(table_sections):
                table_info = table_sections[i]
                self.fill_single_language_table(ws, pair_data, table_info)

    def find_table_sections(self, ws):
        """Находит все секции с таблицами в шаблоне"""
        table_sections = []

        # Ищем заголовки таблиц (например, "Китайский", или цветные ячейки)
        for row_num in range(1, ws.max_row + 1):
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row_num, col_num)

                # Проверяем, является ли ячейка заголовком таблицы
                if self.is_table_header(cell):
                    # Находим границы таблицы
                    table_info = self.get_table_boundaries(ws, row_num, col_num)
                    if table_info:
                        table_sections.append(table_info)

        return table_sections

    def is_table_header(self, cell):
        """Определяет, является ли ячейка заголовком таблицы"""
        if not cell.value:
            return False

        cell_value = str(cell.value).strip()

        # Проверяем различные признаки заголовка таблицы
        indicators = [
            "Китайский", "Английский", "Немецкий", "Французский",
            "→", "Перевод", "Локализация", "Название работы"
        ]

        # Или ячейка с цветной заливкой
        has_fill = cell.fill and cell.fill.start_color and cell.fill.start_color.index != '00000000'

        # Или жирный шрифт
        is_bold = cell.font and cell.font.bold

        return any(indicator in cell_value for indicator in indicators) or has_fill or is_bold

    def get_table_boundaries(self, ws, header_row, header_col):
        """Определяет границы таблицы начиная от заголовка"""

        # Ищем строку с заголовками колонок (Название работы, Тип, Ед-ца, и т.д.)
        column_headers_row = None
        for row in range(header_row + 1, min(header_row + 5, ws.max_row + 1)):
            cell_a = ws.cell(row, 1)
            if cell_a.value and ("Название работы" in str(cell_a.value) or "Параметр" in str(cell_a.value)):
                column_headers_row = row
                break

        if not column_headers_row:
            return None

        # Определяем последнюю строку таблицы (до пустой строки или следующего заголовка)
        last_data_row = column_headers_row
        for row in range(column_headers_row + 1, ws.max_row + 1):
            cell_a = ws.cell(row, 1)

            # Если дошли до промежуточной суммы или пустой строки
            if (not cell_a.value or
                    "Промежуточная сумма" in str(cell_a.value) or
                    "КОНЕЧНАЯ СТОИМОСТЬ" in str(cell_a.value)):
                break

            # Если дошли до следующего заголовка таблицы
            if self.is_table_header(cell_a):
                break

            last_data_row = row

        return {
            'header_row': header_row,
            'column_headers_row': column_headers_row,
            'first_data_row': column_headers_row + 1,
            'last_data_row': last_data_row,
            'subtotal_row': last_data_row + 1
        }

    def duplicate_table_sections(self, ws, existing_sections, needed_count):
        """Дублирует таблицы если языковых пар больше чем таблиц"""

        if not existing_sections or len(existing_sections) >= needed_count:
            return

        # Берем последнюю таблицу как образец для дублирования
        template_section = existing_sections[-1]

        # Вычисляем размер одной таблицы
        table_height = template_section['subtotal_row'] - template_section['header_row'] + 2

        # Дублируем таблицу нужное количество раз
        for i in range(len(existing_sections), needed_count):
            new_start_row = template_section['subtotal_row'] + 2 + (i - len(existing_sections) + 1) * table_height

            # Копируем строки таблицы
            for offset in range(table_height):
                source_row = template_section['header_row'] + offset
                target_row = new_start_row + offset

                # Копируем каждую ячейку в строке
                for col in range(1, 6):  # A-E колонки
                    source_cell = ws.cell(source_row, col)
                    target_cell = ws.cell(target_row, col)

                    # Копируем значение, стиль и форматирование
                    target_cell.value = source_cell.value
                    if source_cell.font:
                        target_cell.font = Font(
                            name=source_cell.font.name,
                            size=source_cell.font.size,
                            bold=source_cell.font.bold,
                            color=source_cell.font.color
                        )
                    if source_cell.fill:
                        target_cell.fill = PatternFill(
                            fill_type=source_cell.fill.fill_type,
                            start_color=source_cell.fill.start_color,
                            end_color=source_cell.fill.end_color
                        )
                    if source_cell.alignment:
                        target_cell.alignment = Alignment(
                            horizontal=source_cell.alignment.horizontal,
                            vertical=source_cell.alignment.vertical
                        )

    def fill_single_language_table(self, ws, pair_data, table_info):
        """Заполняет одну таблицу данными языковой пары"""

        # Обновляем заголовок таблицы
        header_cell = ws.cell(table_info['header_row'], 1)
        header_cell.value = pair_data['pair_name']

        # Очищаем существующие данные в таблице
        for row in range(table_info['first_data_row'], table_info['last_data_row'] + 1):
            for col in range(1, 6):
                ws.cell(row, col).value = None

        # Заполняем новыми данными
        current_row = table_info['first_data_row']
        total_sum = 0

        for service_name, service_data in pair_data.get('services', {}).items():
            for row_data in service_data:
                if row_data['volume'] > 0 or row_data['rate'] > 0:  # Только непустые строки
                    if current_row <= table_info['last_data_row']:
                        ws.cell(current_row, 1).value = row_data['parameter']  # Название работы
                        ws.cell(current_row, 2).value = "Слово"  # Тип
                        ws.cell(current_row, 3).value = row_data['volume']  # Ед-ца
                        ws.cell(current_row, 4).value = row_data['rate']  # Ставка
                        ws.cell(current_row, 5).value = row_data['total']  # Итого

                        total_sum += row_data['total']
                        current_row += 1

        # Обновляем промежуточную сумму
        subtotal_cell = ws.cell(table_info['subtotal_row'], 5)
        subtotal_cell.value = f"{total_sum:.2f}р."

    def fill_additional_services_data(self, ws, project_data: Dict[str, Any]):
        """Заполняет дополнительные услуги в существующие разделы шаблона"""

        additional_services = project_data.get('additional_services', {})
        if not additional_services:
            return

        # Ищем раздел "Подготовка проекта" или другие разделы для дополнительных услуг
        prep_section = self.find_preparation_section(ws)

        if prep_section:
            self.fill_preparation_section(ws, prep_section, additional_services)

        # Обновляем итоговую сумму
        self.update_final_total(ws, project_data)

    def find_preparation_section(self, ws):
        """Находит раздел 'Подготовка проекта' или аналогичный"""

        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row_num, 1)
            if cell.value and "Подготовка проекта" in str(cell.value):

                # Находим строку с заголовками
                headers_row = None
                for r in range(row_num + 1, min(row_num + 5, ws.max_row + 1)):
                    if ws.cell(r, 1).value and "Название работы" in str(ws.cell(r, 1).value):
                        headers_row = r
                        break

                if headers_row:
                    return {
                        'header_row': row_num,
                        'column_headers_row': headers_row,
                        'first_data_row': headers_row + 1,
                        'last_data_row': headers_row + 10  # Предполагаем до 10 строк
                    }

        return None

    def fill_preparation_section(self, ws, section, additional_services):
        """Заполняет раздел подготовки проекта дополнительными услугами"""

        current_row = section['first_data_row']

        # Очищаем существующие данные
        for row in range(section['first_data_row'], section['last_data_row'] + 1):
            for col in range(1, 6):
                ws.cell(row, col).value = None

        # Заполняем дополнительными услугами
        for service_name, service_data in additional_services.items():
            for row_data in service_data:
                if row_data['volume'] > 0 or row_data['rate'] > 0:
                    if current_row <= section['last_data_row']:
                        ws.cell(current_row, 1).value = f"{service_name}: {row_data['parameter']}"
                        ws.cell(current_row, 2).value = "Час"  # или другая единица
                        ws.cell(current_row, 3).value = row_data['volume']
                        ws.cell(current_row, 4).value = row_data['rate']
                        ws.cell(current_row, 5).value = row_data['total']
                        current_row += 1

    def update_final_total(self, ws, project_data):
        """Обновляет итоговую сумму в шаблоне"""

        total_sum = self.calculate_total_from_data(project_data)

        # Ищем ячейку с "КОНЕЧНАЯ СТОИМОСТЬ"
        for row_num in range(1, ws.max_row + 1):
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row_num, col_num)
                if cell.value and "КОНЕЧНАЯ СТОИМОСТЬ" in str(cell.value).upper():
                    # Обновляем ячейку справа или в колонке E
                    result_cell = ws.cell(row_num, 5)  # Колонка E
                    result_cell.value = f"{total_sum:.2f}р."
                    return

    def calculate_total_from_data(self, project_data: Dict[str, Any]) -> float:
        """Вычисляет общую сумму из данных проекта"""
        total = 0.0

        # Языковые пары
        for pair_data in project_data.get('language_pairs', []):
            for service_data in pair_data.get('services', {}).values():
                for row_data in service_data:
                    total += row_data['total']

        # Дополнительные услуги
        for service_data in project_data.get('additional_services', {}).values():
            for row_data in service_data:
                total += row_data['total']

        return total

    def write_header(self, ws, project_data: Dict[str, Any], start_row: int):
        """Записывает заголовок документа"""
        ws[f'A{start_row}'] = f"КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ"
        ws[f'A{start_row}'].font = Font(bold=True, size=16)

        ws[f'A{start_row + 1}'] = f"Проект: {project_data.get('project_name', '')}"
        ws[f'A{start_row + 2}'] = f"Номер/код: {project_data.get('project_code', '')}"
        ws[f'A{start_row + 3}'] = f"Клиент: {project_data.get('client_name', '')}"
        ws[f'A{start_row + 4}'] = f"Дата: {datetime.now().strftime('%d.%m.%Y')}"

    def write_language_pair(self, ws, pair_data: Dict[str, Any], start_row: int) -> int:
        """Записывает данные языковой пары"""
        current_row = start_row

        # Заголовок языковой пары
        ws[f'A{current_row}'] = pair_data['pair_name']
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        current_row += 1

        # Заголовки таблицы
        headers = ["Параметр", "Объем", "Ставка (руб)", "Сумма (руб)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(current_row, col, header)
            cell.font = Font(bold=True)
        current_row += 1

        # Записываем данные услуг
        for service_name, service_data in pair_data.get('services', {}).items():
            # Заголовок услуги
            ws[f'A{current_row}'] = service_name.upper()
            ws[f'A{current_row}'].font = Font(bold=True)
            current_row += 1

            # Данные услуги
            for row_data in service_data:
                ws[f'A{current_row}'] = row_data['parameter']
                ws[f'B{current_row}'] = row_data['volume']
                ws[f'C{current_row}'] = row_data['rate']
                ws[f'D{current_row}'] = row_data['total']
                current_row += 1

            current_row += 1  # Пустая строка после услуги

        return current_row

    def write_additional_services(self, ws, services_data: Dict[str, Any], start_row: int) -> int:
        """Записывает дополнительные услуги"""
        current_row = start_row

        ws[f'A{current_row}'] = "ДОПОЛНИТЕЛЬНЫЕ УСЛУГИ"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        current_row += 2

        for service_name, service_data in services_data.items():
            # Заголовок услуги
            ws[f'A{current_row}'] = service_name
            ws[f'A{current_row}'].font = Font(bold=True)
            current_row += 1

            # Заголовки таблицы
            headers = ["Параметр", "Объем", "Ставка (руб)", "Сумма (руб)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(current_row, col, header)
                cell.font = Font(bold=True)
            current_row += 1

            # Данные услуги
            for row_data in service_data:
                ws[f'A{current_row}'] = row_data['parameter']
                ws[f'B{current_row}'] = row_data['volume']
                ws[f'C{current_row}'] = row_data['rate']
                ws[f'D{current_row}'] = row_data['total']
                current_row += 1

            current_row += 1  # Пустая строка после услуги

        return current_row


class TranslationCostCalculator(QMainWindow):
    """Главное окно приложения"""

    def __init__(self):
        super().__init__()
        self.language_pairs = {}  # словарь виджетов языковых пар
        self.setup_ui()
        self.setup_style()

    def setup_ui(self):
        self.setWindowTitle("Калькулятор стоимости переводческих проектов")
        self.setGeometry(100, 100, 1200, 800)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QHBoxLayout()

        # Создаем Splitter для изменяемого размера панелей
        splitter = QSplitter(Qt.Horizontal)

        # Левая панель - основные данные и управление
        left_panel = self.create_left_panel()
        splitter.addWidget(left_panel)

        # Правая панель - языковые пары и услуги
        right_panel = self.create_right_panel()
        splitter.addWidget(right_panel)

        # Устанавливаем пропорции
        splitter.setSizes([300, 900])

        main_layout.addWidget(splitter)
        central_widget.setLayout(main_layout)

    def create_left_panel(self) -> QWidget:
        """Создает левую панель с основными данными"""
        widget = QWidget()
        layout = QVBoxLayout()

        # Основная информация о проекте
        project_group = QGroupBox("Информация о проекте")
        project_layout = QVBoxLayout()

        # Название проекта
        project_layout.addWidget(QLabel("Название проекта:"))
        self.project_name_edit = QLineEdit()
        project_layout.addWidget(self.project_name_edit)

        # Номер проекта
        project_layout.addWidget(QLabel("Номер/код проекта:"))
        self.project_code_edit = QLineEdit()
        project_layout.addWidget(self.project_code_edit)

        # Клиент
        project_layout.addWidget(QLabel("Название клиента:"))
        self.client_name_edit = QLineEdit()
        project_layout.addWidget(self.client_name_edit)

        # Контактное лицо
        project_layout.addWidget(QLabel("Контактное лицо:"))
        self.contact_person_edit = QLineEdit()
        project_layout.addWidget(self.contact_person_edit)

        # Email
        project_layout.addWidget(QLabel("E-mail:"))
        self.email_edit = QLineEdit()
        project_layout.addWidget(self.email_edit)

        project_group.setLayout(project_layout)
        layout.addWidget(project_group)

        # Управление языковыми парами
        pairs_group = QGroupBox("Языковые пары")
        pairs_layout = QVBoxLayout()

        # Поля для добавления новой пары
        add_pair_layout = QHBoxLayout()

        self.source_lang_edit = QLineEdit()
        self.source_lang_edit.setPlaceholderText("Исходный язык")
        add_pair_layout.addWidget(self.source_lang_edit)

        add_pair_layout.addWidget(QLabel("→"))

        self.target_lang_edit = QLineEdit()
        self.target_lang_edit.setPlaceholderText("Целевой язык")
        add_pair_layout.addWidget(self.target_lang_edit)

        pairs_layout.addLayout(add_pair_layout)

        # Кнопка добавления пары
        self.add_pair_btn = QPushButton("Добавить языковую пару")
        self.add_pair_btn.clicked.connect(self.add_language_pair)
        pairs_layout.addWidget(self.add_pair_btn)

        # Список текущих пар
        pairs_layout.addWidget(QLabel("Текущие пары:"))
        self.pairs_list = QTextEdit()
        self.pairs_list.setMaximumHeight(100)
        self.pairs_list.setReadOnly(True)
        pairs_layout.addWidget(self.pairs_list)

        pairs_group.setLayout(pairs_layout)
        layout.addWidget(pairs_group)

        # Кнопки действий
        actions_group = QGroupBox("Действия")
        actions_layout = QVBoxLayout()

        # Выбор шаблона
        template_layout = QHBoxLayout()
        self.template_path_edit = QLineEdit()
        self.template_path_edit.setPlaceholderText("Путь к шаблону Excel")
        template_layout.addWidget(self.template_path_edit)

        select_template_btn = QPushButton("Выбрать")
        select_template_btn.clicked.connect(self.select_template)
        template_layout.addWidget(select_template_btn)

        actions_layout.addLayout(template_layout)

        # Сохранение
        self.save_excel_btn = QPushButton("Сохранить Excel")
        self.save_excel_btn.clicked.connect(self.save_excel)
        actions_layout.addWidget(self.save_excel_btn)

        self.save_pdf_btn = QPushButton("Сохранить PDF")
        self.save_pdf_btn.clicked.connect(self.save_pdf)
        actions_layout.addWidget(self.save_pdf_btn)

        # Загрузка/сохранение проекта
        actions_layout.addWidget(QFrame())  # разделитель

        self.save_project_btn = QPushButton("Сохранить проект")
        self.save_project_btn.clicked.connect(self.save_project)
        actions_layout.addWidget(self.save_project_btn)

        self.load_project_btn = QPushButton("Загрузить проект")
        self.load_project_btn.clicked.connect(self.load_project)
        actions_layout.addWidget(self.load_project_btn)

        actions_group.setLayout(actions_layout)
        layout.addWidget(actions_group)

        layout.addStretch()  # Добавляем растяжение внизу
        widget.setLayout(layout)
        return widget

    def create_right_panel(self) -> QWidget:
        """Создает правую панель с услугами"""
        widget = QWidget()
        layout = QVBoxLayout()

        # Создаем вкладки
        self.tabs = QTabWidget()

        # Вкладка для языковых пар
        self.pairs_scroll = QScrollArea()
        self.pairs_widget = QWidget()
        self.pairs_layout = QVBoxLayout()
        self.pairs_widget.setLayout(self.pairs_layout)
        self.pairs_scroll.setWidget(self.pairs_widget)
        self.pairs_scroll.setWidgetResizable(True)

        self.tabs.addTab(self.pairs_scroll, "Языковые пары")

        # Вкладка для дополнительных услуг
        self.additional_services_widget = AdditionalServicesWidget()
        additional_scroll = QScrollArea()
        additional_scroll.setWidget(self.additional_services_widget)
        additional_scroll.setWidgetResizable(True)

        self.tabs.addTab(additional_scroll, "Дополнительные услуги")

        layout.addWidget(self.tabs)
        widget.setLayout(layout)
        return widget

    def setup_style(self):
        """Настраивает стили приложения"""
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QGroupBox {
                font-weight: bold;
                border: 1px solid #cccccc;
                border-radius: 5px;
                margin-top: 1ex;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
            QPushButton {
                background-color: #4CAF50;
                border: none;
                color: white;
                padding: 8px 16px;
                text-align: center;
                font-size: 14px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
            QLineEdit, QTextEdit {
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 5px;
            }
            QTableWidget {
                border: 1px solid #ddd;
                border-radius: 4px;
            }
        """)

    def add_language_pair(self):
        """Добавляет новую языковую пару"""
        source_lang = self.source_lang_edit.text().strip()
        target_lang = self.target_lang_edit.text().strip()

        if not source_lang or not target_lang:
            QMessageBox.warning(self, "Ошибка", "Введите оба языка")
            return

        pair_name = f"{source_lang} → {target_lang}"

        if pair_name in self.language_pairs:
            QMessageBox.warning(self, "Ошибка", "Такая языковая пара уже существует")
            return

        # Создаем виджет для языковой пары
        pair_widget = LanguagePairWidget(pair_name)
        self.language_pairs[pair_name] = pair_widget

        # Добавляем в интерфейс
        self.pairs_layout.addWidget(pair_widget)

        # Обновляем список пар
        self.update_pairs_list()

        # Очищаем поля ввода
        self.source_lang_edit.clear()
        self.target_lang_edit.clear()

    def update_pairs_list(self):
        """Обновляет список языковых пар"""
        pairs_text = "\n".join(self.language_pairs.keys())
        self.pairs_list.setText(pairs_text)

    def select_template(self):
        """Выбирает файл шаблона Excel"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выбрать шаблон Excel", "",
            "Excel files (*.xlsx *.xls)"
        )
        if file_path:
            self.template_path_edit.setText(file_path)

    def collect_project_data(self) -> Dict[str, Any]:
        """Собирает все данные проекта"""
        data = {
            "project_name": self.project_name_edit.text(),
            "project_code": self.project_code_edit.text(),
            "client_name": self.client_name_edit.text(),
            "contact_person": self.contact_person_edit.text(),
            "email": self.email_edit.text(),
            "language_pairs": [],
            "additional_services": {}
        }

        # Собираем данные языковых пар
        for pair_widget in self.language_pairs.values():
            pair_data = pair_widget.get_data()
            if pair_data["services"]:  # Только если есть выбранные услуги
                data["language_pairs"].append(pair_data)

        # Собираем данные дополнительных услуг
        additional_data = self.additional_services_widget.get_data()
        if additional_data:
            data["additional_services"] = additional_data

        return data

    def save_excel(self):
        """Сохраняет данные в Excel файл"""
        if not self.client_name_edit.text().strip():
            QMessageBox.warning(self, "Ошибка", "Введите название клиента")
            return

        project_data = self.collect_project_data()

        if not project_data["language_pairs"] and not project_data["additional_services"]:
            QMessageBox.warning(self, "Ошибка", "Добавьте хотя бы одну услугу")
            return

        # Генерируем имя файла
        client_name = project_data["client_name"].replace(" ", "_")
        date_str = datetime.now().strftime("%Y%m%d")
        filename = f"КП_{client_name}_{date_str}.xlsx"

        # Выбираем папку для сохранения
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить Excel файл", filename,
            "Excel files (*.xlsx)"
        )

        if not file_path:
            return

        # Экспортируем данные
        template_path = self.template_path_edit.text().strip()
        exporter = ExcelExporter(template_path if template_path else None)

        if exporter.export_to_excel(project_data, file_path):
            QMessageBox.information(self, "Успех", f"Файл сохранен: {file_path}")
        else:
            QMessageBox.critical(self, "Ошибка", "Не удалось сохранить файл")

    def save_pdf(self):
        """Сохраняет данные в PDF файл"""
        QMessageBox.information(self, "Функция PDF",
                                "Функция сохранения в PDF будет реализована в следующей версии.\n"
                                "Пока вы можете сохранить Excel файл и конвертировать его в PDF вручную.")

    def save_project(self):
        """Сохраняет проект в JSON файл"""
        if not self.project_name_edit.text().strip():
            QMessageBox.warning(self, "Ошибка", "Введите название проекта")
            return

        project_data = self.collect_project_data()

        # Генерируем имя файла
        project_name = project_data["project_name"].replace(" ", "_")
        filename = f"Проект_{project_name}.json"

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить проект", filename,
            "JSON files (*.json)"
        )

        if not file_path:
            return

        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(project_data, f, ensure_ascii=False, indent=2)
            QMessageBox.information(self, "Успех", f"Проект сохранен: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить проект: {e}")

    def load_project(self):
        """Загружает проект из JSON файла"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Загрузить проект", "",
            "JSON files (*.json)"
        )

        if not file_path:
            return

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                project_data = json.load(f)

            self.load_project_data(project_data)
            QMessageBox.information(self, "Успех", "Проект загружен")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить проект: {e}")

    def load_project_data(self, project_data: Dict[str, Any]):
        """Загружает данные проекта в интерфейс"""
        # Основная информация
        self.project_name_edit.setText(project_data.get("project_name", ""))
        self.project_code_edit.setText(project_data.get("project_code", ""))
        self.client_name_edit.setText(project_data.get("client_name", ""))
        self.contact_person_edit.setText(project_data.get("contact_person", ""))
        self.email_edit.setText(project_data.get("email", ""))

        # Очищаем текущие языковые пары
        for pair_widget in self.language_pairs.values():
            pair_widget.setParent(None)
        self.language_pairs.clear()

        # Загружаем языковые пары
        for pair_data in project_data.get("language_pairs", []):
            pair_name = pair_data["pair_name"]

            # Создаем виджет пары
            pair_widget = LanguagePairWidget(pair_name)
            self.language_pairs[pair_name] = pair_widget
            self.pairs_layout.addWidget(pair_widget)

            # Загружаем данные услуг
            services = pair_data.get("services", {})

            if "translation" in services:
                pair_widget.translation_group.setChecked(True)
                self.load_table_data(pair_widget.translation_group.table, services["translation"])

            if "editing" in services:
                pair_widget.editing_group.setChecked(True)
                self.load_table_data(pair_widget.editing_group.table, services["editing"])

        # Обновляем список пар
        self.update_pairs_list()

        # Загружаем дополнительные услуги
        additional_services = project_data.get("additional_services", {})
        for service_name, service_data in additional_services.items():
            if service_name in self.additional_services_widget.service_groups:
                group = self.additional_services_widget.service_groups[service_name]
                group.setChecked(True)
                self.load_table_data(group.table, service_data)

    def load_table_data(self, table: QTableWidget, data: List[Dict[str, Any]]):
        """Загружает данные в таблицу"""
        for row, row_data in enumerate(data):
            if row < table.rowCount():
                table.item(row, 1).setText(str(row_data.get("volume", 0)))
                table.item(row, 2).setText(str(row_data.get("rate", 0)))
                table.item(row, 3).setText(str(row_data.get("total", 0)))

    def calculate_total_cost(self) -> float:
        """Вычисляет общую стоимость проекта"""
        total = 0.0

        # Суммируем языковые пары
        for pair_widget in self.language_pairs.values():
            pair_data = pair_widget.get_data()
            for service_data in pair_data.get("services", {}).values():
                for row_data in service_data:
                    total += row_data["total"]

        # Суммируем дополнительные услуги
        additional_data = self.additional_services_widget.get_data()
        for service_data in additional_data.values():
            for row_data in service_data:
                total += row_data["total"]

        return total


def main():
    """Основная функция запуска приложения"""
    app = QApplication(sys.argv)

    # Настраиваем иконку приложения (если есть)
    # app.setWindowIcon(QIcon("icon.png"))

    # Создаем и показываем главное окно
    window = TranslationCostCalculator()
    window.show()

    # Запускаем главный цикл приложения
    sys.exit(app.exec())


if __name__ == "__main__":
    main()