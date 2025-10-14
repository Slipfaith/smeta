"""Excel export helpers used by the legacy rate tabs."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, Mapping

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PySide6.QtWidgets import QTableWidget


def table_to_df(table: QTableWidget) -> pd.DataFrame:
    """Convert a :class:`~PySide6.QtWidgets.QTableWidget` into a DataFrame."""

    headers: list[str] = []
    for column in range(table.columnCount()):
        header_item = table.horizontalHeaderItem(column)
        headers.append(header_item.text() if header_item else f"Column {column + 1}")

    rows: list[list[str]] = []
    for row in range(table.rowCount()):
        row_values: list[str] = []
        for column in range(table.columnCount()):
            item = table.item(row, column)
            row_values.append(item.text() if item else "")
        rows.append(row_values)

    return pd.DataFrame(rows, columns=headers)


def apply_excel_styles(
    worksheet,
    dataframe: pd.DataFrame,
    numeric_columns: Iterable[str],
    *,
    number_formats: Mapping[str, str] | None = None,
) -> None:
    """Apply simple formatting to ``worksheet`` to improve readability."""

    numeric_columns = [col for col in numeric_columns if col in dataframe.columns]
    numeric_indices = [dataframe.columns.get_loc(col) for col in numeric_columns]
    number_formats = dict(number_formats or {})

    row_count = dataframe.shape[0]
    column_count = dataframe.shape[1]

    for column_index in range(1, column_count + 1):
        cell = worksheet.cell(row=1, column=column_index)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")

    for column_index, column_name in enumerate(dataframe.columns, start=1):
        column_letter = get_column_letter(column_index)
        if (column_index - 1) in numeric_indices:
            format_string = number_formats.get(column_name, "#,##0")
            max_width = len(column_name)
            for row_index in range(2, row_count + 2):
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.number_format = format_string
                cell.alignment = Alignment(horizontal="right", vertical="center")
                value = cell.value
                if value is not None:
                    max_width = max(max_width, len(str(value)))
            worksheet.column_dimensions[column_letter].width = max(10, max_width + 2)
        else:
            max_width = len(column_name)
            for row_index in range(2, row_count + 2):
                value = worksheet.cell(row=row_index, column=column_index).value
                if value is not None:
                    max_width = max(max_width, len(str(value)))
            worksheet.column_dimensions[column_letter].width = max_width + 2


def export_rate_tables(sheet_dfs: Mapping[str, pd.DataFrame], out_path: str | Path) -> None:
    """Export rate tables to an Excel workbook."""

    out_path = Path(out_path)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, dataframe in sheet_dfs.items():
            frame = dataframe.copy()
            numeric_columns = list(frame.columns[2:])
            for column in numeric_columns:
                frame[column] = pd.to_numeric(frame[column], errors="coerce")

            frame.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]

            try:
                currency = sheet_name.split("_")[1]
            except IndexError:
                currency = "USD"

            symbol_map = {"USD": "$", "EUR": "€", "RUB": "₽", "CNY": "¥"}
            decimals_map = {"USD": 3, "EUR": 3, "RUB": 2, "CNY": 2}
            symbol = symbol_map.get(currency, "$")
            decimals = decimals_map.get(currency, 2)
            decimal_part = "0" * decimals

            number_formats = {
                "Basic": f'"{symbol}"#,##0.{decimal_part}' if decimals else f'"{symbol}"#,##0',
                "Complex": f'"{symbol}"#,##0.{decimal_part}' if decimals else f'"{symbol}"#,##0',
                "Hour": f'"{symbol}"#,##0',
            }

            apply_excel_styles(
                worksheet,
                frame,
                numeric_columns,
                number_formats=number_formats,
            )


def export_logtab(lang_totals_table: QTableWidget, file_stats_table: QTableWidget, out_path: str | Path) -> None:
    """Export the LogTab tables into an Excel workbook."""

    out_path = Path(out_path)

    lang_df = table_to_df(lang_totals_table)
    for column in lang_df.columns[1:]:
        lang_df[column] = pd.to_numeric(lang_df[column], errors="coerce")

    file_df = table_to_df(file_stats_table)
    for column in file_df.columns[2:]:
        file_df[column] = pd.to_numeric(file_df[column], errors="coerce")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        lang_df.to_excel(writer, sheet_name="Language Totals", index=False)
        file_df.to_excel(writer, sheet_name="File-Level Statistics", index=False)

        lang_ws = writer.sheets["Language Totals"]
        apply_excel_styles(lang_ws, lang_df, lang_df.columns[1:])

        file_ws = writer.sheets["File-Level Statistics"]
        apply_excel_styles(file_ws, file_df, file_df.columns[2:])


def export_memoqtab(
    memoq_table: QTableWidget,
    out_path: str | Path,
    error_messages: Iterable[str],
) -> None:
    """Export the MemoQ table and optional error log."""

    if memoq_table.rowCount() == 0:
        return

    out_path = Path(out_path)
    dataframe = table_to_df(memoq_table)
    for column in dataframe.columns[1:]:
        dataframe[column] = (
            pd.to_numeric(dataframe[column], errors="coerce")
            .fillna(0)
            .astype(int)
        )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, sheet_name="MemoQ Stats", index=False)
        worksheet = writer.sheets["MemoQ Stats"]

        row_count = dataframe.shape[0]
        column_count = dataframe.shape[1]

        for column_index in range(1, column_count + 1):
            cell = worksheet.cell(row=1, column=column_index)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")

        for column_index in range(2, column_count + 1):
            for row_index in range(2, row_count + 2):
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")

        for column_index in range(1, column_count + 1):
            column_letter = get_column_letter(column_index)
            max_length = 0
            for row_index in range(1, row_count + 2):
                value = worksheet.cell(row=row_index, column=column_index).value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            worksheet.column_dimensions[column_letter].width = max_length + 2

    error_messages = list(error_messages)
    if error_messages:
        log_path = out_path.with_name(out_path.stem + "_errors.log")
        with log_path.open("w", encoding="utf-8") as handle:
            for message in error_messages:
                handle.write(f"{message}\n")
