from __future__ import annotations

from openpyxl import Workbook

from logic.excel_exporter import (
    ExcelExporter,
    ProjectSetupRenderer,
    PS_HDR,
    PS_START_PH,
    PS_END_PH,
)


def _build_project_setup_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    ws["A1"] = PS_START_PH
    ws["A2"] = PS_HDR["param"]
    ws["B2"] = PS_HDR["unit"]
    ws["C2"] = PS_HDR["qty"]
    ws["D2"] = PS_HDR["rate"]
    ws["E2"] = PS_HDR["total"]
    ws["A3"] = None
    ws["A4"] = PS_END_PH

    return wb


def test_project_setup_renderer_uses_english_discount_and_markup(tmp_path):
    wb = _build_project_setup_template()
    ws = wb.active

    exporter = ExcelExporter(
        template_path=str(tmp_path / "template.xlsx"),
        currency="USD",
        log_path=str(tmp_path / "export.log"),
        lang="en",
    )

    renderer = ProjectSetupRenderer(exporter)
    renderer.render(
        ws,
        1,
        {
            "rows": [
                {
                    "parameter": "Setup",
                    "unit": "час",
                    "volume": 2,
                    "rate": 100,
                }
            ],
            "discount_percent": 10,
            "markup_percent": 5,
        },
    )

    values = {
        cell
        for row in ws.iter_rows(min_row=1, max_row=8, values_only=True)
        for cell in row
        if isinstance(cell, str) and cell
    }

    assert "Project setup fee" in values
    assert "Discount 10%" in values
    assert "Markup 5%" in values
