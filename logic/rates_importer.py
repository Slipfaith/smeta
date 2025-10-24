"""Excel rates importer and language pair matching utilities."""

from __future__ import annotations

import logging
import os
import re
from dataclasses import dataclass
from typing import Dict, Iterable, IO, List, Optional, Tuple, Union

import langcodes
import openpyxl

from .xml_parser_common import expand_language_code, language_identity

logger = logging.getLogger(__name__)

RateRecord = Dict[str, float]
RatesMap = Dict[Tuple[str, str], RateRecord]


def _normalize_language(name: str) -> str:
    """Return a normalized ISO language code for *name*.

    The helper first tries :func:`logic.xml_parser_common.language_identity`
    so that script or territory hints from the Excel sheet are preserved.
    When the value cannot be resolved that way, we fall back to the
    previous heuristics based on :func:`langcodes.standardize_tag` and
    :func:`langcodes.find`, finally returning the lower-cased input if all
    lookups fail.
    """
    language, script, territory = language_identity(name)
    if language:
        parts = [language]
        if script:
            parts.append(script.lower())
        if territory:
            parts.append(territory.lower())
        return "-".join(parts)

    cleaned = re.sub(r"\s*\(.*?\)", "", name)
    cleaned = re.split(r"[,/\-]", cleaned, maxsplit=1)[0].strip()
    try:
        tag = langcodes.standardize_tag(cleaned)
    except Exception:
        tag = ""
    if tag:
        return tag.replace("_", "-").lower()

    try:
        return langcodes.find(cleaned).language
    except LookupError:
        return cleaned.lower()


def _language_name(code: str) -> str:
    """Return English display name for a language *code*."""
    if not code:
        return ""

    normalized = code.replace("_", "-")
    try:
        tag = langcodes.standardize_tag(normalized)
    except Exception:
        tag = normalized

    pretty = expand_language_code(tag, locale="en")
    if pretty:
        return pretty

    try:
        return langcodes.Language.get(tag).display_name("en")
    except Exception:
        try:
            return langcodes.Language.make(tag).display_name("en")
        except Exception:
            return code


def load_rates_from_excel(
    path_or_stream: Union[str, os.PathLike[str], IO[bytes]],
    currency: str,
    rate_type: str,
) -> RatesMap:
    """Load rates from *path* for the given *currency* and *rate_type*.

    Parameters
    ----------
    path_or_stream:
        Path to the Excel workbook or a binary stream containing the workbook contents.
    currency:
        Currency code selected in the GUI (e.g. ``"USD"``).
    rate_type:
        Rate table version, such as ``"R1"`` or ``"R2"``.

    Returns
    -------
    dict
        Mapping of ``(src_lang, tgt_lang)`` pairs to rate dictionaries
        with keys ``basic``, ``complex`` and ``hour``.
    """
    sheet_name = f"{rate_type}_{currency}".upper()
    logger.debug(
        "Loading Excel rates: sheet=%s, currency=%s, rate_type=%s, path_type=%s",
        sheet_name,
        currency,
        rate_type,
        type(path_or_stream).__name__,
    )

    stream = path_or_stream
    if hasattr(stream, "seek"):
        logger.debug("Resetting stream pointer before reading workbook")
        stream.seek(0)
    wb = openpyxl.load_workbook(stream, data_only=True)
    logger.debug("Workbook sheets available: %s", ", ".join(wb.sheetnames))
    if sheet_name not in wb.sheetnames:
        logger.error("Requested sheet %s not found in workbook", sheet_name)
        raise ValueError(f"Sheet {sheet_name} not found in workbook")

    ws = wb[sheet_name]
    rates: RatesMap = {}
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        src, tgt, basic, complex_, hour = row
        logger.debug(
            "Row %d raw values: src=%r, tgt=%r, basic=%r, complex=%r, hour=%r",
            index,
            src,
            tgt,
            basic,
            complex_,
            hour,
        )
        if not src or not tgt:
            logger.debug("Row %d skipped: missing source or target", index)
            continue
        src_code = _normalize_language(str(src))
        tgt_code = _normalize_language(str(tgt))
        rate_record = {
            "basic": float(basic or 0),
            "complex": float(complex_ or 0),
            "hour": float(hour or 0),
        }
        logger.debug(
            "Row %d normalized: src_code=%s, tgt_code=%s, rates=%s",
            index,
            src_code,
            tgt_code,
            rate_record,
        )
        rates[(src_code, tgt_code)] = rate_record

    logger.debug("Loaded %d rate entries from sheet %s", len(rates), sheet_name)
    return rates


@dataclass
class PairMatch:
    """Result of matching a GUI language pair with Excel rates."""

    gui_source: str
    gui_target: str
    excel_source: str
    excel_target: str
    rates: Optional[RateRecord]


def match_pairs(
    gui_pairs: Iterable[Tuple[str, str]],
    rates: RatesMap,
    manual_codes: Optional[Dict[Tuple[str, str], Tuple[str, str]]] = None,
    manual_names: Optional[Dict[Tuple[str, str], Tuple[str, str]]] = None,
) -> List[PairMatch]:
    """Match GUI language pairs against loaded *rates*.

    Parameters
    ----------
    gui_pairs:
        Iterable of ``(source, target)`` language names as entered in the GUI.
    rates:
        Rates map produced by :func:`load_rates_from_excel`.
    manual_codes:
        Optional mapping of GUI pairs to explicit ``(src_code, tgt_code)`` tuples
        chosen by the user.  When provided, the codes are used as-is without
        applying language normalization.
    manual_names:
        Optional mapping of GUI pairs to the Excel display names chosen by the
        user.  When present, these names are returned in the resulting
        :class:`PairMatch` regardless of whether a rate is found.

    Returns
    -------
    list[PairMatch]
        A list describing how each GUI pair was matched.  If a corresponding
        rate was not found, ``rates`` attribute will be ``None``.
    """
    gui_pairs = list(gui_pairs)
    results: List[PairMatch] = []
    available_sources = {src for src, _ in rates}
    available_targets = {tgt for _, tgt in rates}
    logger.debug(
        "Matching %d GUI pairs against %d available rates", len(gui_pairs), len(rates)
    )
    sources_display = ", ".join(sorted(available_sources)) or "<none>"
    targets_display = ", ".join(sorted(available_targets)) or "<none>"
    logger.debug("Available source codes: %s", sources_display)
    logger.debug("Available target codes: %s", targets_display)
    manual_codes = manual_codes or {}
    manual_names = manual_names or {}
    manual_pairs = set(manual_codes) | set(manual_names)

    for gui_src, gui_tgt in gui_pairs:
        key = (gui_src, gui_tgt)
        if key in manual_pairs:
            override_codes = manual_codes.get(key)
            override_names = manual_names.get(key, ("", ""))
            src_code = override_codes[0] if override_codes else ""
            tgt_code = override_codes[1] if override_codes else ""
            rate = rates.get((src_code, tgt_code)) if override_codes else None
            excel_src = override_names[0] or (
                _language_name(src_code) if src_code in available_sources else ""
            )
            excel_tgt = override_names[1] or (
                _language_name(tgt_code) if tgt_code in available_targets else ""
            )
        else:
            src_code = _normalize_language(gui_src)
            tgt_code = _normalize_language(gui_tgt)
            rate = rates.get((src_code, tgt_code))
            excel_src = _language_name(src_code) if src_code in available_sources else ""
            excel_tgt = _language_name(tgt_code) if tgt_code in available_targets else ""
        logger.debug(
            "Pair matched: gui=(%s, %s) -> codes=(%s, %s), found_rate=%s, excel_names=(%s, %s)",
            gui_src,
            gui_tgt,
            src_code,
            tgt_code,
            rate,
            excel_src,
            excel_tgt,
        )
        results.append(
            PairMatch(
                gui_source=gui_src,
                gui_target=gui_tgt,
                excel_source=excel_src,
                excel_target=excel_tgt,
                rates=rate,
            )
        )
    return results
