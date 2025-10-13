# logic/excel_exporter.py
import os
import logging
import re
import textwrap
import threading
from typing import Dict, Any, List, Optional, Tuple, Callable
from copy import deepcopy
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Protection
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor
from openpyxl.drawing.image import Image
from resource_utils import resource_path

from .service_config import ServiceConfig
from .translation_config import tr
from .excel_process import apply_separators

CURRENCY_SYMBOLS = {"RUB": "₽", "EUR": "€", "USD": "$"}

# Шаблон теперь ищем относительно корня проекта или временной папки PyInstaller
DEFAULT_TEMPLATE_PATH = resource_path("templates/шаблон.modern.xlsx")

START_PH = "{{translation_table}}"
END_PH = "{{subtotal_translation_table}}"

HDR = {
    "param": "{{taskname}}",
    "type": "{{type}}",
    "unit": "{{unit}}",
    "qty": "{{quantity}}",
    "rate": "{{rate}}",
    "total": "{{total_translation_table}}",
}
HDR_TITLES = {
    HDR["param"]: "Название работы",
    HDR["type"]: "Тип",
    HDR["unit"]: "Ед-ца",
    HDR["qty"]: "Кол-во",
    HDR["rate"]: "Ставка",
    HDR["total"]: "Итого",
}

# блок запуска и управления проектом
PS_START_PH = "{{project_setup}}"
PS_END_PH = "{{subtotal_project_setup}}"
PS_HDR = {
    "param": "{{taskname.project_setup}}",
    "unit": "{{unit.project_setup}}",
    "qty": "{{quantity.project_setup}}",
    "rate": "{{rate.project_setup}}",
    "total": "{{total_project_setup_table}}",
}
PS_HDR_TITLES = {
    PS_HDR["param"]: "Названия работ",
    PS_HDR["unit"]: "Ед-ца",
    PS_HDR["qty"]: "Кол-во",
    PS_HDR["rate"]: "Ставка",
    PS_HDR["total"]: "Итого",
}

# блок дополнительных услуг
ADD_START_PH = "{{add.service_table}}"
ADD_END_PH = "{{subtotal.add.service}}"
ADD_HDR = {
    "param": "{{add.service.taskname}}",
    "unit": "{{add.service.unit}}",
    "qty": "{{add.service.quantity}}",
    "rate": "{{add.service.rate}}",
    "total": "{{add.service.table}}",
}
ADD_HDR_TITLES = {
    ADD_HDR["param"]: "Параметр",
    ADD_HDR["unit"]: "Ед-ца",
    ADD_HDR["qty"]: "Кол-во",
    ADD_HDR["rate"]: "Ставка",
    ADD_HDR["total"]: "Итого",
}


# ----------------------------- БЛОКИ РЕНДЕРИНГА -----------------------------


class BlockRenderer:
    """Базовый рендерер таблиц.

    Подготовка данных реализуется в наследниках, а вставка в лист и
    заполнение значений делегируется общему методу ``_render_block``
    экспортера.
    """

    def __init__(
        self,
        exporter: "ExcelExporter",
        start_ph: str,
        end_ph: str,
        hdr_tokens: Dict[str, str],
        titles: Dict[str, str],
    ) -> None:
        self.exporter = exporter
        self.start_ph = start_ph
        self.end_ph = end_ph
        self.hdr_tokens = hdr_tokens
        self.titles = titles

    def template_sheet(self, ws: Worksheet) -> Worksheet:  # pragma: no cover - simple
        return ws

    def header_title(self, data: Any) -> str:  # pragma: no cover - override
        return ""

    def prepare_items(self, data: Any) -> List[Dict[str, Any]]:  # pragma: no cover - override
        raise NotImplementedError

    def discount_percent(self, data: Any) -> float:  # pragma: no cover - override
        return 0.0

    def markup_percent(self, data: Any) -> float:  # pragma: no cover - override
        return 0.0

    def render(
        self, ws: Worksheet, start_row: int, data: Any
    ) -> Tuple[int, Optional[str], Optional[str], Optional[str]]:
        template_ws = self.template_sheet(ws)
        items = self.prepare_items(data)
        title = self.header_title(data)
        discount = self.discount_percent(data)
        markup = self.markup_percent(data)
        return self.exporter._render_block(
            ws,
            template_ws,
            start_row,
            self.start_ph,
            self.end_ph,
            self.hdr_tokens,
            items,
            self.titles,
            title,
            discount_percent=discount,
            markup_percent=markup,
        )


class TranslationBlockRenderer(BlockRenderer):
    def __init__(self, exporter: "ExcelExporter") -> None:
        super().__init__(exporter, START_PH, END_PH, HDR, exporter.hdr_titles)

    def template_sheet(self, ws: Worksheet) -> Worksheet:
        wb = ws.parent
        return wb["Languages"] if "Languages" in wb.sheetnames else ws

    def header_title(self, pair: Dict[str, Any]) -> str:
        return pair.get("pair_name") or pair.get("header_title") or ""

    def prepare_items(self, pair: Dict[str, Any]) -> List[Dict[str, Any]]:
        translation_rows = (pair.get("services") or {}).get("translation", [])
        deleted_keys = set()
        deleted_names = set()
        filtered_rows: List[Dict[str, Any]] = []
        for row in translation_rows:
            if row.get("deleted"):
                key = row.get("key")
                if key is not None:
                    deleted_keys.add(key)
                name = row.get("name")
                if name:
                    deleted_names.add(str(name))
                continue
            filtered_rows.append(row)
        translation_rows = filtered_rows
        only_new_mode = bool(
            pair.get("only_new_repeats")
            or pair.get("only_new_repeats_mode")
        )
        if only_new_mode:
            items: List[Dict[str, Any]] = []
            for row in translation_rows:
                items.append(
                    {
                        "parameter": row.get("parameter")
                        or tr(row.get("name"), self.exporter.lang),
                        "unit": tr("Слово", self.exporter.lang),
                        "volume": self.exporter._to_number(row.get("volume") or 0),
                        "rate": self.exporter._to_number(row.get("rate") or 0),
                        "multiplier": row.get("multiplier"),
                        "is_base": bool(row.get("is_base")),
                    }
                )
            return items

        data_map: Dict[str, Dict[str, Any]] = {}
        extras: List[Dict[str, Any]] = []
        default_keys = {cfg.get("key") for cfg in ServiceConfig.TRANSLATION_ROWS}
        for row in translation_rows:
            key = row.get("key") or row.get("name")
            lname = str(key).lower() if key else ""
            if "new" in lname or "нов" in lname:
                std_key = "new"
            elif "95" in lname:
                std_key = "fuzzy_95_99"
            elif "75" in lname:
                std_key = "fuzzy_75_94"
            elif "100" in lname:
                std_key = "reps_100_30"
            else:
                std_key = key
            if std_key in default_keys:
                if std_key in data_map:
                    extras.append(row)
                else:
                    data_map[std_key] = row
            else:
                extras.append(row)

        items: List[Dict[str, Any]] = []
        for cfg in ServiceConfig.TRANSLATION_ROWS:
            key = cfg.get("key")
            if key in deleted_keys or cfg.get("name") in deleted_names:
                continue
            src = data_map.get(key, {})
            items.append(
                {
                    "parameter": src.get("parameter")
                    or tr(cfg.get("name"), self.exporter.lang),
                    "unit": tr("Слово", self.exporter.lang),
                    "volume": self.exporter._to_number(src.get("volume") or 0),
                    "rate": self.exporter._to_number(src.get("rate") or 0),
                    "multiplier": src.get("multiplier", cfg.get("multiplier")),
                    "is_base": bool(src.get("is_base", cfg.get("is_base"))),
                }
            )

        for row in extras:
            items.append(
                {
                    "parameter": row.get("parameter")
                    or tr(row.get("name"), self.exporter.lang),
                    "unit": tr("Слово", self.exporter.lang),
                    "volume": self.exporter._to_number(row.get("volume") or 0),
                    "rate": self.exporter._to_number(row.get("rate") or 0),
                    "multiplier": row.get("multiplier"),
                    "is_base": bool(row.get("is_base")),
                }
            )

        return items

    def discount_percent(self, pair: Dict[str, Any]) -> float:
        return float(pair.get("discount_percent", 0.0) or 0.0)

    def markup_percent(self, pair: Dict[str, Any]) -> float:
        return float(pair.get("markup_percent", 0.0) or 0.0)


class ProjectSetupRenderer(BlockRenderer):
    def __init__(self, exporter: "ExcelExporter") -> None:
        super().__init__(exporter, PS_START_PH, PS_END_PH, PS_HDR, exporter.ps_hdr_titles)

    def template_sheet(self, ws: Worksheet) -> Worksheet:
        wb = ws.parent
        if "Setupfee" in wb.sheetnames:
            return wb["Setupfee"]
        if "ProjectSetup" in wb.sheetnames:
            return wb["ProjectSetup"]
        return ws

    def header_title(self, data: Any) -> str:
        return tr("Запуск и управление проектом", self.exporter.lang)

    def prepare_items(self, data: Any) -> List[Dict[str, Any]]:
        if isinstance(data, dict):
            items_list = data.get("rows", [])
        else:
            items_list = data
        items: List[Dict[str, Any]] = []
        for it in items_list:
            items.append(
                {
                    "parameter": it.get("parameter", ""),
                    "unit": tr("час", self.exporter.lang),
                    "volume": self.exporter._to_number(it.get("volume", 0)),
                    "rate": self.exporter._to_number(it.get("rate", 0)),
                }
            )
        return items

    def discount_percent(self, data: Any) -> float:
        if isinstance(data, dict):
            return float(data.get("discount_percent", 0.0) or 0.0)
        return 0.0

    def markup_percent(self, data: Any) -> float:
        if isinstance(data, dict):
            return float(data.get("markup_percent", 0.0) or 0.0)
        return 0.0


class AdditionalServicesRenderer(BlockRenderer):
    def __init__(self, exporter: "ExcelExporter") -> None:
        super().__init__(exporter, ADD_START_PH, ADD_END_PH, ADD_HDR, exporter.add_hdr_titles)

    def template_sheet(self, ws: Worksheet) -> Worksheet:
        wb = ws.parent
        if "Addservice" in wb.sheetnames:
            return wb["Addservice"]
        if "AdditionalServices" in wb.sheetnames:
            return wb["AdditionalServices"]
        return ws

    def header_title(self, block: Dict[str, Any]) -> str:
        return tr(block.get("header_title", "Дополнительные услуги"), self.exporter.lang)

    def prepare_items(self, block: Dict[str, Any]) -> List[Dict[str, Any]]:
        items: List[Dict[str, Any]] = []
        for it in block.get("rows", []):
            unit = it.get("unit", "")
            items.append(
                {
                    "parameter": it.get("parameter", ""),
                    "unit": tr(unit, self.exporter.lang) if unit else "",
                    "volume": self.exporter._to_number(it.get("volume", 0)),
                    "rate": self.exporter._to_number(it.get("rate", 0)),
                }
            )
        return items

    def discount_percent(self, block: Dict[str, Any]) -> float:
        return float(block.get("discount_percent", 0.0) or 0.0)

    def markup_percent(self, block: Dict[str, Any]) -> float:
        return float(block.get("markup_percent", 0.0) or 0.0)


class ExcelExporter:
    """Экспорт проектных данных по блоку {{translation_table}} … {{subtotal_translation_table}}."""

    def __init__(
        self,
        template_path: Optional[str] = None,
        currency: str = "RUB",
        lang: str = "ru",
    ):
        self.template_path = template_path or DEFAULT_TEMPLATE_PATH
        self.currency = currency
        self.currency_symbol = CURRENCY_SYMBOLS.get(currency, "")
        self.rate_fmt = self._currency_format(3)
        self.total_fmt = self._currency_format(2)
        self.lang = lang
        self.hdr_titles = {k: tr(v, lang) for k, v in HDR_TITLES.items()}
        self.ps_hdr_titles = {k: tr(v, lang) for k, v in PS_HDR_TITLES.items()}
        self.add_hdr_titles = {k: tr(v, lang) for k, v in ADD_HDR_TITLES.items()}
        self.subtotal_title = f"{tr('Промежуточная сумма', lang)} ({currency}):"
        self.discount_title = tr("Скидка", lang)
        self.markup_title = tr("Наценка", lang)
        self.logger = logging.getLogger("ExcelExporter")
        self.logger.setLevel(logging.DEBUG)
        self.logger.debug(
            "Initialized ExcelExporter with template %s", self.template_path
        )

    def _currency_format(self, decimals: int) -> str:
        sym = self.currency_symbol
        if self.currency == "USD":
            return f'"{sym}"#,##0.{"0"*decimals}'
        return f'#,##0.{"0"*decimals} "{sym}"'

    def _replace_currency_formats(self, wb: Workbook) -> None:
        """Replace dollar signs in cell number formats with the current symbol.

        Excel stores some currency formats using the ``[$...-...]`` syntax. A
        naive string replacement could break such patterns and lead to a broken
        workbook that Excel tries to recover.  Here we first substitute the
        whole bracketed expression with the current currency symbol wrapped in
        quotes and only then replace any remaining ``$`` characters.
        """
        if self.currency == "USD":
            return

        bracket_re = re.compile(r"\[\$[^-]*-[^\]]*\]")

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    fmt = cell.number_format
                    if not isinstance(fmt, str) or "$" not in fmt:
                        continue
                    fmt = bracket_re.sub(f'"{self.currency_symbol}"', fmt)
                    if "$" in fmt:
                        fmt = fmt.replace("$", self.currency_symbol)
                    cell.number_format = fmt

    def _to_number(self, value: Any) -> Any:
        if isinstance(value, str):
            try:
                return float(value.replace(",", "."))
            except ValueError:
                return value
        return value

    def _apply_rate_format(self, cell: Cell) -> None:
        """Apply correct number format for rate cells.

        If the rate has no fractional part, use two decimals instead of three.
        """
        val = cell.value
        num: Optional[float] = None
        if isinstance(val, str):
            if val.startswith("="):
                m = re.match(r"=([A-Z]+\d+)\*([0-9.,]+)$", val)
                if m:
                    ref, mult = m.groups()
                    try:
                        base = cell.parent[ref].value
                        if isinstance(base, str):
                            base = float(base.replace(",", "."))
                        num = float(base) * float(mult.replace(",", "."))
                    except Exception:
                        num = None
                else:
                    num = None
            else:
                try:
                    num = float(val.replace(",", "."))
                except ValueError:
                    num = None
        elif isinstance(val, (int, float)):
            num = float(val)

        if num is not None and not (isinstance(val, str) and val.startswith("=")):
            cell.value = num
        if num is not None:
            if num.is_integer():
                cell.number_format = self.total_fmt
                return
            scaled = round(num * 1000)
            if scaled % 10 == 0:
                cell.number_format = self.total_fmt
                return
        cell.number_format = self.rate_fmt

    def _set_print_area(self, ws: Worksheet) -> None:
        last_col = get_column_letter(ws.max_column)
        ws.print_area = f"A1:{last_col}{ws.max_row}"

    def _fit_sheet_to_page(self, ws: Worksheet) -> None:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        self._set_print_area(ws)

    def _fix_trailing_zeroes(self, wb: Workbook) -> None:
        """Заменяет три нуля после разделителя на два нуля во всех ячейках."""

        pattern = re.compile(r"(\d+[,.]\d*?)000\b")
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, (int, float)):
                        try:
                            num = float(value)
                        except Exception:
                            continue
                        scaled = round(num * 1000)
                        if scaled % 10 == 0:
                            fmt = cell.number_format or ""
                            # Skip cells explicitly marked as general numbers (e.g. quantity columns)
                            if fmt == "General":
                                continue
                            if "000" in fmt:
                                cell.number_format = fmt.replace("000", "00", 1)
                            else:
                                cell.number_format = self.total_fmt
                    elif isinstance(value, str):
                        new_val = pattern.sub(r"\1" + "00", value)
                        if new_val != value:
                            cell.value = new_val

    def _get_cell_width(self, ws: Worksheet, cell: Cell) -> float:
        """Calculate width in characters for a cell, respecting merged ranges."""
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                width = 0.0
                for col in range(merged.min_col, merged.max_col + 1):
                    letter = get_column_letter(col)
                    width += ws.column_dimensions[letter].width or 8.43
                return width
        letter = get_column_letter(cell.column)
        return ws.column_dimensions[letter].width or 8.43

    def _adjust_row_height_for_text(self, ws: Worksheet, cell: Cell) -> None:
        """Adjust row height to fit wrapped text inside a (possibly merged) cell."""
        text = str(cell.value or "")
        if not text:
            return
        width = self._get_cell_width(ws, cell)
        if width <= 0:
            return
        max_chars = max(1, int(width))
        lines = 0
        for part in text.splitlines() or [text]:
            wrapped = textwrap.wrap(part, width=max_chars) or [""]
            lines += len(wrapped)
        row_dim = ws.row_dimensions[cell.row]
        base_height = row_dim.height or ws.sheet_format.defaultRowHeight or 15
        row_dim.height = base_height * max(1, lines)

    # ----------------------------- ПУБЛИЧНЫЙ АПИ -----------------------------

    def export_to_excel(
        self,
        project_data: Dict[str, Any],
        output_path: str,
        fit_to_page: bool = False,
        progress_callback: Optional[Callable[[int, str], None]] = None,
    ) -> bool:
        """Export project data to an Excel file.

        Parameters
        ----------
        project_data: Dict[str, Any]
            Исходные данные проекта.
        output_path: str
            Путь к итоговому файлу.
        fit_to_page: bool, optional
            Подгонять ли лист под страницу при печати.
        progress_callback: Callable[[int, str], None], optional
            Функция для обновления прогресса.
        """
        try:
            self.logger.info("Starting export to %s", output_path)
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Шаблон не найден: {self.template_path}")

            pairs_count = len(project_data.get("language_pairs", []))
            add_count = len(project_data.get("additional_services", []))
            total_steps = 7 + pairs_count + add_count
            progress = 0

            def step(message: str = "") -> None:
                nonlocal progress
                progress += 1
                if progress_callback:
                    percent = int(progress / total_steps * 100)
                    progress_callback(percent, message)

            if progress_callback:
                progress_callback(0, "Загрузка шаблона")

            wb = load_workbook(self.template_path)
            self._replace_currency_formats(wb)
            step("Шаблон загружен")

            quotation_ws = (
                wb["Quotation"] if "Quotation" in wb.sheetnames else wb.active
            )

            subtot_cells: List[str] = []
            discount_cells: List[str] = []
            markup_cells: List[str] = []
            current_row = 13

            last_row, ps_cell, ps_discount, ps_markup = self._render_project_setup_table(
                quotation_ws, project_data, current_row
            )
            if ps_cell:
                subtot_cells.append(ps_cell)
            if ps_discount:
                discount_cells.append(ps_discount)
            if ps_markup:
                markup_cells.append(ps_markup)
            step("Настройка проекта")
            current_row = last_row + 1

            last_row, tr_cells, tr_discount_cells, tr_markup_cells = self._render_translation_blocks(
                quotation_ws,
                project_data,
                current_row,
                progress_callback=lambda name: step(f"Перевод {name}"),
            )
            subtot_cells += tr_cells
            discount_cells += tr_discount_cells
            markup_cells += tr_markup_cells
            current_row = last_row + 1

            (
                last_row,
                add_cells,
                add_discount_cells,
                add_markup_cells,
            ) = self._render_additional_services_tables(
                quotation_ws,
                project_data,
                current_row,
                progress_callback=lambda name: step(f"{name}"),
            )
            subtot_cells += add_cells
            discount_cells += add_discount_cells
            markup_cells += add_markup_cells

            self.logger.debug("Subtotal cells collected: %s", subtot_cells)

            for name in (
                "ProjectSetup",
                "Setupfee",
                "Languages",
                "AdditionalServices",
                "Addservice",
            ):
                if name in wb.sheetnames:
                    del wb[name]

            step("Удаление шаблонных листов")

            self._fill_text_placeholders(
                quotation_ws,
                project_data,
                subtot_cells,
                discount_cells,
                markup_cells,
                start_row=1,
                wb=wb,
            )
            step("Заполнение плейсхолдеров")

            if "Vat" in wb.sheetnames:
                del wb["Vat"]

            if fit_to_page:
                self._fit_sheet_to_page(quotation_ws)
                step("Подгонка страницы")
            else:
                self._set_print_area(quotation_ws)
                step("Установка области печати")

            self._fix_trailing_zeroes(wb)

            step("Добавление логотипа")
            template = project_data.get("legal_entity")
            if template:
                self._insert_logo(quotation_ws, template)

            self.logger.info("Saving workbook to %s", output_path)
            step("Сохранение файла")
            wb.save(output_path)
            self._apply_separators_async(output_path)

            if progress_callback:
                progress_callback(100, "Готово")

            self.logger.info("Export completed successfully")
            return True
        except Exception:
            self.logger.exception("Ошибка экспорта в Excel")
            return False

    def _apply_separators_async(self, output_path: str) -> None:
        """Adjust number separators without blocking the UI."""

        def worker() -> None:
            try:
                success = apply_separators(output_path, self.lang)
                if not success:
                    self.logger.debug(
                        "Skipping separator adjustment for %s", output_path
                    )
            except Exception:
                # ``apply_separators`` already guards against most failures, but we
                # still protect the background thread from propagating exceptions
                # back to the caller.
                self.logger.exception(
                    "Unexpected error while adjusting separators for %s", output_path
                )

        thread = threading.Thread(
            target=worker,
            name="excel-separators",
            daemon=True,
        )
        thread.start()

    # ----------------------------- ПОИСК/КОПИРОВАНИЕ -----------------------------

    def _find_first(
        self, ws: Worksheet, token: str, row_from: int = 1
    ) -> Optional[Tuple[int, int]]:
        self.logger.debug("Searching for '%s' starting from row %d", token, row_from)
        for r in range(row_from, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip() == token:
                    self.logger.debug("  found '%s' at row %d col %d", token, r, c)
                    return r, c
        self.logger.debug("  token '%s' not found", token)
        return None

    def _find_below(
        self, ws: Worksheet, start_row: int, token: str
    ) -> Optional[Tuple[int, int]]:
        self.logger.debug("Searching below row %d for '%s'", start_row, token)
        for r in range(start_row + 1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip() == token:
                    self.logger.debug("  found '%s' at row %d col %d", token, r, c)
                    return r, c
        self.logger.debug("  token '%s' not found below row %d", token, start_row)
        return None

    def _copy_style(self, s: Cell, d: Cell) -> None:
        d.value = s.value
        if s.font:
            d.font = Font(
                name=s.font.name,
                size=s.font.size,
                bold=s.font.bold,
                italic=s.font.italic,
                vertAlign=s.font.vertAlign,
                underline=s.font.underline,
                strike=s.font.strike,
                color=s.font.color,
            )
        if s.alignment:
            d.alignment = Alignment(
                horizontal=s.alignment.horizontal,
                vertical=s.alignment.vertical,
                text_rotation=s.alignment.text_rotation,
                wrap_text=s.alignment.wrap_text,
                shrink_to_fit=s.alignment.shrink_to_fit,
                indent=s.alignment.indent,
            )
        if s.fill:
            d.fill = PatternFill(
                fill_type=s.fill.fill_type,
                start_color=s.fill.start_color,
                end_color=s.fill.end_color,
            )
        if s.border:
            d.border = Border(
                left=s.border.left,
                right=s.border.right,
                top=s.border.top,
                bottom=s.border.bottom,
                diagonal=s.border.diagonal,
                diagonalUp=s.border.diagonalUp,
                diagonalDown=s.border.diagonalDown,
                outline=s.border.outline,
                vertical=s.border.vertical,
                horizontal=s.border.horizontal,
            )
        if s.has_style and s.number_format:
            d.number_format = s.number_format
        if s.protection:
            d.protection = Protection(
                locked=s.protection.locked, hidden=s.protection.hidden
            )

    def _copy_block(
        self,
        dst_ws: Worksheet,
        src_ws: Worksheet,
        src_start: int,
        src_end: int,
        dst_start: int,
    ) -> None:
        """Копирует строки [src_start..src_end] со страницы src_ws на позицию dst_start в dst_ws
        со стилями и слияниями. Используется, чтобы копировать неизменённый шаблонный
        блок даже после того, как первый блок уже заполнен данными."""
        height = src_end - src_start + 1
        self.logger.debug(
            "- Copying rows %d-%d from template to position %d-%d",
            src_start,
            src_end,
            dst_start,
            dst_start + height - 1,
        )
        for i in range(height):
            sr = src_start + i
            dr = dst_start + i
            self.logger.debug("  copying row %d -> %d", sr, dr)
            try:
                dst_ws.row_dimensions[dr].height = src_ws.row_dimensions[sr].height
            except Exception:
                pass
            for c in range(1, src_ws.max_column + 1):
                self._copy_style(src_ws.cell(sr, c), dst_ws.cell(dr, c))
        merges = []
        for m in src_ws.merged_cells.ranges:
            sr, sc, er, ec = m.min_row, m.min_col, m.max_row, m.max_col
            if src_start <= sr and er <= src_end:
                delta = dst_start - src_start
                merges.append((sr + delta, sc, er + delta, ec))
        for sr, sc, er, ec in merges:
            ref = f"{get_column_letter(sc)}{sr}:{get_column_letter(ec)}{er}"
            # Убираем перекрывающиеся с уже существующими слияния, чтобы Excel
            # не ругался на повреждённые записи при открытии файла.
            overlap = []
            for m in dst_ws.merged_cells.ranges:
                if not (
                    m.max_row < sr or m.min_row > er or m.max_col < sc or m.min_col > ec
                ):
                    overlap.append(str(m))
            for mref in overlap:
                try:
                    dst_ws.unmerge_cells(mref)
                except Exception:
                    pass
            try:
                dst_ws.merge_cells(ref)
            except Exception:
                pass

        # Копирование изображений, закреплённых в указанном диапазоне
        delta = dst_start - src_start
        for img in getattr(src_ws, "_images", []):
            anchor = getattr(img, "anchor", None)
            if anchor is None:
                continue
            if isinstance(anchor, TwoCellAnchor):
                from_attr = "from_" if hasattr(anchor, "from_") else "_from"
                start_row = getattr(anchor, from_attr).row + 1
                end_row = anchor.to.row + 1
                if src_start <= start_row and end_row <= src_end:
                    new_anchor = deepcopy(anchor)
                    getattr(new_anchor, from_attr).row += delta
                    new_anchor.to.row += delta
                    new_img = Image(img._data())
                    new_img.anchor = new_anchor
                    dst_ws.add_image(new_img)
            elif isinstance(anchor, OneCellAnchor):
                row = anchor._from.row + 1
                if src_start <= row <= src_end:
                    new_anchor = deepcopy(anchor)
                    new_anchor._from.row += delta
                    new_img = Image(img._data())
                    new_img.anchor = new_anchor
                    dst_ws.add_image(new_img)

    def _insert_logo(self, ws: Worksheet, template: str) -> None:
        """Insert company logo for the specified template into the worksheet.

        Parameters
        ----------
        ws: Worksheet
            Лист Excel, на который вставляется логотип.
        template: str
            Код шаблона, соответствующий имени файла логотипа.
        """
        logo_path = resource_path(f"templates/logos/{template}.png")
        if not os.path.exists(logo_path):
            self.logger.debug("Logo not found for template %s: %s", template, logo_path)
            return
        try:
            img = Image(logo_path)
            ws.add_image(img, "A1")
        except Exception as e:
            self.logger.debug("Failed to insert logo %s: %s", logo_path, e)

    def _shift_images(self, ws: Worksheet, idx: int, amount: int) -> None:
        """Сдвигает изображения, расположенные на листе, если были вставлены строки."""
        for image in getattr(ws, "_images", []):
            anchor = getattr(image, "anchor", None)
            if anchor is None:
                continue
            if isinstance(anchor, TwoCellAnchor):
                from_attr = "from_" if hasattr(anchor, "from_") else "_from"
                if getattr(anchor, from_attr).row >= idx - 1:
                    getattr(anchor, from_attr).row += amount
                    anchor.to.row += amount
            elif isinstance(anchor, OneCellAnchor):
                if anchor._from.row >= idx - 1:
                    anchor._from.row += amount

    def _insert_rows(self, ws: Worksheet, idx: int, amount: int) -> None:
        """Вставляет строки и корректирует позиции изображений."""
        ws.insert_rows(idx, amount)
        self._shift_images(ws, idx, amount)

    # ----------------------------- МАП КОЛОНОК -----------------------------

    def _header_map(
        self,
        ws: Worksheet,
        headers_row: int,
        hdr_tokens: Dict[str, str] = HDR,
        default_mapping: Optional[Dict[str, int]] = None,
    ) -> Dict[str, int]:
        mapping: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(headers_row, c).value
            if isinstance(v, str):
                t = v.strip()
                for key, tok in hdr_tokens.items():
                    if t == tok:
                        mapping[key] = c
        if default_mapping:
            for k, v in default_mapping.items():
                mapping.setdefault(k, v)
        if mapping:
            self.logger.debug("Header map at row %d: %s", headers_row, mapping)
            return mapping
        mapping = {
            "param": 1,
            "type": 2,
            "unit": 3,
            "qty": 4,
            "rate": 5,
            "total": 6,
        }
        self.logger.debug(
            "Header tokens not found at row %d, using default mapping %s",
            headers_row,
            mapping,
        )
        return mapping

    # ----------------------------- ОСНОВНОЙ РЕНДЕР -----------------------------

    def _render_block(
        self,
        ws: Worksheet,
        template_ws: Worksheet,
        start_row: int,
        start_ph: str,
        end_ph: str,
        hdr_tokens: Dict[str, str],
        rows: List[Dict[str, Any]],
        titles: Dict[str, str],
        header_title: Optional[str] = None,
        discount_percent: float = 0.0,
        markup_percent: float = 0.0,
    ) -> Tuple[int, Optional[str], Optional[str], Optional[str]]:
        tpl_start = self._find_first(template_ws, start_ph)
        if not tpl_start:
            return start_row - 1, None
        tpl_start_row, _ = tpl_start
        tpl_end = self._find_below(template_ws, tpl_start_row, end_ph)
        if not tpl_end:
            return start_row - 1, None
        tpl_end_row, _ = tpl_end

        template_height = tpl_end_row - tpl_start_row + 1
        headers_rel = 1
        first_data_rel = 2
        subtotal_rel = template_height - 1

        self._insert_rows(ws, start_row, template_height)
        self._copy_block(ws, template_ws, tpl_start_row, tpl_end_row, start_row)

        extra_start = self._find_first(ws, start_ph, start_row + template_height)
        if extra_start:
            extra_end = self._find_below(ws, extra_start[0], end_ph)
            if extra_end:
                ws.delete_rows(extra_start[0], extra_end[0] - extra_start[0] + 1)

        block_top = start_row
        t_headers_row = block_top + headers_rel
        t_first_data = block_top + first_data_rel
        t_subtotal_row = block_top + subtotal_rel

        if header_title:
            for c in range(1, ws.max_column + 1):
                if ws.cell(block_top, c).value == start_ph:
                    ws.cell(block_top, c, header_title)
                    break

        default_map = {k: i + 1 for i, k in enumerate(hdr_tokens.keys())}
        hmap = self._header_map(ws, t_headers_row, hdr_tokens, default_map)
        first_col = min(hmap.values())
        last_col = max(hmap.values())

        for c in range(first_col, last_col + 1):
            v = ws.cell(t_headers_row, c).value
            if isinstance(v, str) and v.strip() in titles:
                ws.cell(t_headers_row, c, titles[v.strip()])

        need = len(rows)
        cur_cap = max(0, t_subtotal_row - t_first_data)
        if need > cur_cap:
            add = need - cur_cap
            self._insert_rows(ws, t_subtotal_row, add)
            tpl_row = t_first_data if cur_cap > 0 else t_subtotal_row - 1
            merges_to_copy = [
                m
                for m in ws.merged_cells.ranges
                if m.min_row == m.max_row == tpl_row
                and first_col <= m.min_col
                and m.max_col <= last_col
            ]
            for k in range(add):
                dst = t_subtotal_row + k
                for c in range(first_col, last_col + 1):
                    self._copy_style(ws.cell(tpl_row, c), ws.cell(dst, c))
                for m in merges_to_copy:
                    sc, ec = m.min_col, m.max_col
                    ref = f"{get_column_letter(sc)}{dst}:{get_column_letter(ec)}{dst}"
                    ws.merge_cells(ref)
            t_subtotal_row += add
        elif need < cur_cap:
            delete_count = cur_cap - need
            if delete_count > 0:
                ws.delete_rows(t_first_data + need, delete_count)
                t_subtotal_row -= delete_count

        for rr in range(t_first_data, t_subtotal_row):
            for c in range(first_col, last_col + 1):
                ws.cell(rr, c).value = None

        for m in list(ws.merged_cells.ranges):
            if (
                m.min_row >= t_headers_row
                and m.max_row < t_subtotal_row
                and m.min_col >= first_col
                and m.max_col <= last_col
            ):
                try:
                    ws.unmerge_cells(str(m))
                except Exception:
                    pass

        col_param = hmap.get("param", first_col)
        col_unit = hmap.get("unit", col_param + 1)
        col_qty = hmap.get("qty", col_unit + 1)
        col_rate = hmap.get("rate", col_qty + 1)
        col_total = hmap.get("total", col_rate + 1)

        header_ref = f"{get_column_letter(first_col)}{block_top}:{get_column_letter(last_col)}{block_top}"
        try:
            ws.unmerge_cells(header_ref)
        except Exception:
            pass
        ws.merge_cells(header_ref)
        if header_title:
            ws.cell(block_top, first_col, header_title)

        r = t_first_data
        row_numbers: List[int] = []
        for it in rows:
            row_numbers.append(r)
            ws.cell(r, col_param, it.get("parameter", ""))
            unit_cell = ws.cell(r, col_unit, it.get("unit", ""))
            if it.get("unit"):
                unit_cell.alignment = Alignment(horizontal="center", vertical="center")
            qty_cell = ws.cell(r, col_qty, self._to_number(it.get("volume", 0)))
            qty_cell.alignment = Alignment(horizontal="right", vertical="top")
            qty_cell.number_format = "General"
            r += 1

        base_rate_cell = None
        qtyL = get_column_letter(col_qty)
        rateL = get_column_letter(col_rate)
        totalL = get_column_letter(col_total)
        for idx, it in enumerate(rows):
            rr = row_numbers[idx]
            if it.get("is_base"):
                cell = ws.cell(rr, col_rate, self._to_number(it.get("rate", 0)))
                base_rate_cell = f"{rateL}{rr}"
            elif base_rate_cell and it.get("multiplier") is not None:
                mult = self._to_number(it.get("multiplier"))
                cell = ws.cell(rr, col_rate, f"={base_rate_cell}*{mult}")
            else:
                if isinstance(it.get("rate"), str) and it.get("rate", "").startswith("="):
                    cell = ws.cell(rr, col_rate, it.get("rate"))
                else:
                    cell = ws.cell(rr, col_rate, self._to_number(it.get("rate", 0)))
            self._apply_rate_format(cell)
            total_cell = ws.cell(rr, col_total, f"={qtyL}{rr}*{rateL}{rr}")
            total_cell.number_format = self.total_fmt

        discount_value = max(0.0, min(100.0, float(discount_percent or 0.0)))
        markup_value = max(0.0, min(100.0, float(markup_percent or 0.0)))
        subtotal_label = self.subtotal_title

        for c in range(first_col, last_col + 1):
            if ws.cell(t_subtotal_row, c).value == end_ph:
                ws.cell(t_subtotal_row, c, subtotal_label)
                break

        subtotal_formula = "0"
        if r != t_first_data:
            subtotal_formula = f"SUM({totalL}{t_first_data}:{totalL}{r - 1})"

        base_formula = subtotal_formula
        discount_cell_ref: Optional[str] = None
        markup_cell_ref: Optional[str] = None
        if discount_value > 0 and base_formula != "0":
            discount_str = format(discount_value, "g")
            discount_formula = f"({base_formula})*{discount_str}/100"
            self._insert_rows(ws, t_subtotal_row, 1)
            for c in range(first_col, last_col + 1):
                self._copy_style(
                    ws.cell(t_subtotal_row + 1, c), ws.cell(t_subtotal_row, c)
                )
                if c not in (col_param, col_total):
                    ws.cell(t_subtotal_row, c, None)
            discount_label = f"{self.discount_title} {discount_str}%"
            ws.cell(t_subtotal_row, col_param, discount_label)
            discount_cell = ws.cell(
                t_subtotal_row,
                col_total,
                f"={discount_formula}",
            )
            discount_cell.number_format = self.total_fmt
            discount_cell_ref = f"{totalL}{t_subtotal_row}"
            t_subtotal_row += 1
            subtotal_formula = f"({base_formula})-({discount_formula})"
        else:
            subtotal_formula = base_formula

        if markup_value > 0 and base_formula != "0":
            markup_str = format(markup_value, "g")
            markup_formula = f"({base_formula})*{markup_str}/100"
            self._insert_rows(ws, t_subtotal_row, 1)
            for c in range(first_col, last_col + 1):
                self._copy_style(
                    ws.cell(t_subtotal_row + 1, c), ws.cell(t_subtotal_row, c)
                )
                if c not in (col_param, col_total):
                    ws.cell(t_subtotal_row, c, None)
            markup_label = f"{self.markup_title} {markup_str}%"
            ws.cell(t_subtotal_row, col_param, markup_label)
            markup_cell = ws.cell(
                t_subtotal_row,
                col_total,
                f"={markup_formula}",
            )
            markup_cell.number_format = self.total_fmt
            markup_cell_ref = f"{totalL}{t_subtotal_row}"
            t_subtotal_row += 1
            subtotal_formula = f"({subtotal_formula})+({markup_formula})"

        subtotal_cell = ws.cell(
            t_subtotal_row,
            col_total,
            f"={subtotal_formula}",
        )
        subtotal_cell.number_format = self.total_fmt

        return (
            t_subtotal_row,
            f"{totalL}{t_subtotal_row}",
            discount_cell_ref,
            markup_cell_ref,
        )

    def _render_translation_blocks(
        self,
        ws: Worksheet,
        project_data: Dict[str, Any],
        start_row: int,
        progress_callback: Optional[Callable[[str], None]] = None,
    ) -> Tuple[int, List[str], List[str], List[str]]:
        pairs: List[Dict[str, Any]] = project_data.get("language_pairs", [])
        if not pairs:
            return start_row - 1, [], [], []
        self.logger.debug("Rendering %d translation pair(s)", len(pairs))

        pairs = sorted(
            pairs,
            key=lambda p: (
                p.get("pair_name", "").split(" - ")[1]
                if " - " in p.get("pair_name", "")
                else p.get("pair_name", "")
            ),
        )

        renderer = TranslationBlockRenderer(self)
        current_row = start_row
        subtot_cells: List[str] = []
        discount_cells: List[str] = []
        markup_cells: List[str] = []
        for pair in pairs:
            last_row, subtotal_cell, discount_cell, markup_cell = renderer.render(
                ws, current_row, pair
            )
            if subtotal_cell:
                subtot_cells.append(subtotal_cell)
            if discount_cell:
                discount_cells.append(discount_cell)
            if markup_cell:
                markup_cells.append(markup_cell)
            current_row = last_row + 1
            if progress_callback:
                progress_callback(pair.get("pair_name", ""))

        return current_row - 1, subtot_cells, discount_cells, markup_cells

    def _render_project_setup_table(
        self, ws: Worksheet, project_data: Dict[str, Any], start_row: int
    ) -> Tuple[int, Optional[str], Optional[str], Optional[str]]:
        setup_data = project_data.get("project_setup", [])
        discount = project_data.get("project_setup_discount_percent", 0.0)
        markup = project_data.get("project_setup_markup_percent", 0.0)
        if isinstance(setup_data, dict):
            items: List[Dict[str, Any]] = setup_data.get("rows", [])
            discount = setup_data.get("discount_percent", discount)
            markup = setup_data.get("markup_percent", markup)
        else:
            items = setup_data
        if not items:
            return start_row - 1, None, None, None
        renderer = ProjectSetupRenderer(self)
        payload = {
            "rows": items,
            "discount_percent": discount,
            "markup_percent": markup,
        }
        last_row, cell, discount_cell, markup_cell = renderer.render(
            ws, start_row, payload
        )
        return last_row, cell, discount_cell, markup_cell

    def _render_additional_services_tables(
        self,
        ws: Worksheet,
        project_data: Dict[str, Any],
        start_row: int,
        progress_callback: Optional[Callable[[str], None]] = None,
    ) -> Tuple[int, List[str], List[str], List[str]]:
        blocks: List[Dict[str, Any]] = project_data.get("additional_services") or []
        if not blocks:
            return start_row - 1, [], [], []
        renderer = AdditionalServicesRenderer(self)
        current_row = start_row
        subtot_cells: List[str] = []
        discount_cells: List[str] = []
        markup_cells: List[str] = []
        for block in blocks:
            last_row, cell, discount_cell, markup_cell = renderer.render(
                ws, current_row, block
            )
            if cell:
                subtot_cells.append(cell)
            if discount_cell:
                discount_cells.append(discount_cell)
            if markup_cell:
                markup_cells.append(markup_cell)
            current_row = last_row + 1
            if progress_callback:
                progress_callback(block.get("header_title", ""))
        return current_row - 1, subtot_cells, discount_cells, markup_cells

    # ----------------------------- ТЕКСТОВЫЕ ПЛЕЙСХОЛДЕРЫ -----------------------------

    def _fill_text_placeholders(
        self,
        ws: Worksheet,
        project_data: Dict[str, Any],
        subtot_cells: List[str],
        discount_cells: List[str],
        markup_cells: List[str],
        start_row: int = 1,
        wb: Optional[Workbook] = None,
    ) -> None:
        # общий итог — именно как формула SUM по найденным субтоталам (чтобы Excel считал сам)
        total_formula = f"=SUM({','.join(subtot_cells)})" if subtot_cells else "0"
        self.logger.debug("Total formula calculated: %s", total_formula)
        discount_formula = (
            f"=SUM({','.join(discount_cells)})" if discount_cells else ""
        )
        markup_formula = (
            f"=SUM({','.join(markup_cells)})" if markup_cells else ""
        )

        # формируем строку всех языков: сорс (первый), затем уникальные таргеты
        source_lang = ""
        target_langs: List[str] = []
        for p in project_data.get("language_pairs", []):
            pair_name = p.get("pair_name", "")
            if "→" in pair_name:
                src, tgt = [s.strip() for s in pair_name.split("→", 1)]
            elif "-" in pair_name:
                src, tgt = [s.strip() for s in pair_name.split("-", 1)]
            else:
                src, tgt = "", pair_name.strip()
            if not source_lang and src:
                source_lang = src
            if tgt:
                target_langs.append(tgt)
        # удаляем дубликаты в порядке появления
        uniq_targets: List[str] = []
        for lang in target_langs:
            if lang and lang not in uniq_targets:
                uniq_targets.append(lang)
        target_langs_str = ", ".join(uniq_targets)

        currency_code = project_data.get("currency", self.currency)

        strict_map = {
            "{{project_name}}": project_data.get("project_name", ""),
            "{{client}}": project_data.get("client_name", ""),
            "{{client_name}}": project_data.get("contact_person", ""),
            "{{client_email}}": project_data.get("email", ""),
            "{{PM_name}}": project_data.get("pm_name", ""),
            "{{PM_email}}": project_data.get("pm_email", ""),
            "{{Entity}}": project_data.get(
                "entity_name", project_data.get("company_name", "")
            ),
            "{{Entity_address}}": project_data.get(
                "entity_address", project_data.get("email", "")
            ),
            "{{target_langs}}": target_langs_str,
            "{{currency}}": self.currency,
        }
        self.logger.debug("Filling text placeholders with map: %s", strict_map)

        total_cell_ref: Optional[str] = None

        for r in range(start_row, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str):
                    t = v.strip()
                    if t == "{{total}}":
                        target_row = r
                        insert_offset = 0
                        if discount_formula:
                            self._insert_rows(ws, r + insert_offset, 1)
                            for c2 in range(1, ws.max_column + 1):
                                self._copy_style(
                                    ws.cell(r + insert_offset + 1, c2),
                                    ws.cell(r + insert_offset, c2),
                                )
                                ws.cell(r + insert_offset, c2).value = None
                            label_col = 1
                            ws.cell(r + insert_offset, label_col, f"{self.discount_title}")
                            discount_cell = ws.cell(
                                r + insert_offset, c, discount_formula
                            )
                            discount_cell.number_format = self.total_fmt
                            for c2 in range(1, ws.max_column + 1):
                                cell2 = ws.cell(r + insert_offset, c2)
                                val2 = cell2.value
                                if isinstance(val2, str) and "{{$}}" in val2:
                                    cell2.value = val2.replace(
                                        "{{$}}", currency_code
                                    )
                            insert_offset += 1
                        if markup_formula:
                            self._insert_rows(ws, r + insert_offset, 1)
                            for c2 in range(1, ws.max_column + 1):
                                self._copy_style(
                                    ws.cell(r + insert_offset + 1, c2),
                                    ws.cell(r + insert_offset, c2),
                                )
                                ws.cell(r + insert_offset, c2).value = None
                            label_col = 1
                            ws.cell(r + insert_offset, label_col, f"{self.markup_title}")
                            markup_cell = ws.cell(
                                r + insert_offset, c, markup_formula
                            )
                            markup_cell.number_format = self.total_fmt
                            for c2 in range(1, ws.max_column + 1):
                                cell2 = ws.cell(r + insert_offset, c2)
                                val2 = cell2.value
                                if isinstance(val2, str) and "{{$}}" in val2:
                                    cell2.value = val2.replace(
                                        "{{$}}", currency_code
                                    )
                            insert_offset += 1
                        target_row = r + insert_offset
                        total_cell = ws.cell(target_row, c, total_formula)
                        total_cell.number_format = self.total_fmt
                        total_cell_ref = total_cell.coordinate
                        # Replace currency code placeholder in the total row
                        for c2 in range(1, ws.max_column + 1):
                            cell2 = ws.cell(target_row, c2)
                            val2 = cell2.value
                            if isinstance(val2, str) and "{{$}}" in val2:
                                cell2.value = val2.replace("{{$}}", currency_code)
                    else:
                        new_v = v
                        replaced_target_langs = False
                        for ph, val in strict_map.items():
                            if ph in new_v:
                                new_v = new_v.replace(ph, str(val))
                                if ph == "{{target_langs}}":
                                    replaced_target_langs = True
                        if new_v != v:
                            cell = ws.cell(r, c, new_v)
                            if replaced_target_langs:
                                cell.alignment = Alignment(
                                    horizontal=cell.alignment.horizontal,
                                    vertical=cell.alignment.vertical,
                                    text_rotation=cell.alignment.text_rotation,
                                    wrap_text=True,
                                    shrink_to_fit=cell.alignment.shrink_to_fit,
                                    indent=cell.alignment.indent,
                                )
                                self._adjust_row_height_for_text(ws, cell)

        if wb is not None and total_cell_ref:
            self._insert_vat_section(wb, ws, total_cell_ref, project_data)

    # ----------------------------- НДС -----------------------------

    def _insert_vat_section(
        self,
        wb,
        ws: Worksheet,
        total_cell_ref: str,
        project_data: Dict[str, Any],
    ) -> None:
        """Copy VAT table from 'Vat' sheet and insert after total row."""
        vat_rate = self._to_number(project_data.get("vat_rate", 0) or 0)
        if not isinstance(vat_rate, (int, float)) or vat_rate <= 0 or "Vat" not in wb.sheetnames:
            return

        vat_ws = wb["Vat"]
        rows = vat_ws.max_row
        cols = vat_ws.max_column
        total_row = ws[total_cell_ref].row

        self._insert_rows(ws, total_row + 1, rows)

        for r in range(rows):
            for c in range(1, cols + 1):
                src = vat_ws.cell(r + 1, c)
                dst = ws.cell(total_row + 1 + r, c)
                self._copy_style(src, dst)

        repl_map = {
            "{{%vat}}": f"{vat_rate}%",
            "{{total_vat}}": f"={total_cell_ref}*{vat_rate}/100",
            "{{total.with_vat}}": f"={total_cell_ref}*{100+vat_rate}/100",
            "{{$}}": self.currency,
        }

        for r in range(total_row + 1, total_row + 1 + rows):
            for c in range(1, cols + 1):
                cell = ws.cell(r, c)
                if not isinstance(cell.value, str):
                    continue
                new_val = cell.value
                had_total = False
                for ph, val in repl_map.items():
                    if ph in new_val:
                        if ph in ("{{total_vat}}", "{{total.with_vat}}"):
                            had_total = True
                            rep = val
                            if new_val.strip() != ph and rep.startswith("="):
                                rep = rep.lstrip("=")
                        else:
                            rep = val
                        new_val = new_val.replace(ph, rep)
                if new_val != cell.value:
                    cell.value = new_val
                if had_total:
                    cell.number_format = self.total_fmt
