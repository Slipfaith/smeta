# services/excel_export.py

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def table_to_df(table):
    """
    Преобразуем QTableWidget в DataFrame.
    """
    columns = []
    for col in range(table.columnCount()):
        header_item = table.horizontalHeaderItem(col)
        if header_item:
            columns.append(header_item.text())
        else:
            columns.append(f"Column {col+1}")

    data = []
    for row in range(table.rowCount()):
        row_data = []
        for col in range(table.columnCount()):
            item = table.item(row, col)
            row_data.append(item.text() if item else "")
        data.append(row_data)

    df = pd.DataFrame(data, columns=columns)
    return df

def apply_excel_styles(ws, df, numeric_cols, num_formats=None):
    """Apply basic formatting to an Excel sheet.

    ``numeric_cols`` lists columns that should be right-aligned and formatted
    according to ``num_formats`` (mapping column name -> Excel number format).
    If a column is not present in ``num_formats`` it falls back to ``#,##0``.
    """

    row_count = df.shape[0]
    col_count = df.shape[1]

    numeric_indices = [df.columns.get_loc(c) for c in numeric_cols if c in df.columns]
    num_formats = num_formats or {}

    # Заголовки
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

    for i, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(i)
        if (i - 1) in numeric_indices:
            max_len_data = 0
            for row_idx in range(row_count):
                val = df.iloc[row_idx, i - 1]
                if pd.notnull(val):
                    txt = str(val)
                    if len(txt) > max_len_data:
                        max_len_data = len(txt)
            fmt = num_formats.get(col_name, '#,##0')
            for row_idx in range(2, row_count + 2):
                cell = ws.cell(row=row_idx, column=i)
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal='right', vertical='center')
            ws.column_dimensions[col_letter].width = max(10, max_len_data + 2)
        else:
            max_length = len(col_name)
            for row_idx in range(row_count):
                val = df.iloc[row_idx, i - 1]
                if pd.notnull(val):
                    length = len(str(val))
                    if length > max_length:
                        max_length = length
            ws.column_dimensions[col_letter].width = max_length + 2


def export_rate_tables(sheet_dfs, out_path):
    """Экспорт нескольких таблиц ставок в Excel.

    sheet_dfs: dict[str, pd.DataFrame] - ключи это имена листов,
    значения - DataFrame со ставками (колонки: Source, Target, Basic, Complex, Hour).
    Для числовых колонок применяется форматирование через ``apply_excel_styles``.
    """
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, df in sheet_dfs.items():
            # Ensure numeric columns are floats so Excel treats them as numbers
            numeric_cols = df.columns[2:]
            for col in numeric_cols:
                df[col] = pd.to_numeric(df[col], errors="coerce")

            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            # Determine currency from sheet name (e.g. "R1_USD")
            try:
                currency = sheet_name.split("_")[1]
            except IndexError:
                currency = "USD"

            symbol_map = {"USD": "$", "EUR": "€", "RUB": "₽", "CNY": "¥"}
            decimals_map = {"USD": 3, "EUR": 3, "RUB": 2, "CNY": 2}
            symbol = symbol_map.get(currency, "$")
            decimals = decimals_map.get(currency, 2)

            dec_str = "0" * decimals
            num_formats = {
                "Basic": f'"{symbol}"#,##0.{dec_str}' if decimals > 0 else f'"{symbol}"#,##0',
                "Complex": f'"{symbol}"#,##0.{dec_str}' if decimals > 0 else f'"{symbol}"#,##0',
                "Hour": f'"{symbol}"#,##0',
            }

            apply_excel_styles(ws, df, numeric_cols, num_formats=num_formats)

def export_logtab(lang_totals_table, file_stats_table, out_path):
    """
    Логика, которая была в LogTab.export_to_excel.
    Экспортируем 2 таблицы (Language Totals, File-Level Statistics) в один Excel.
    """
    df_lang = table_to_df(lang_totals_table)
    df_file = table_to_df(file_stats_table)

    numeric_cols_lang = df_lang.columns[1:]
    for col in numeric_cols_lang:
        df_lang[col] = pd.to_numeric(df_lang[col], errors='coerce')

    numeric_cols_file = df_file.columns[2:]
    for col in numeric_cols_file:
        df_file[col] = pd.to_numeric(df_file[col], errors='coerce')

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df_lang.to_excel(writer, sheet_name='Language Totals', index=False)
        df_file.to_excel(writer, sheet_name='File-Level Statistics', index=False)

        sheet_lang = writer.sheets['Language Totals']
        apply_excel_styles(sheet_lang, df_lang, numeric_cols_lang)

        sheet_file = writer.sheets['File-Level Statistics']
        apply_excel_styles(sheet_file, df_file, numeric_cols_file)

def export_memoqtab(memoq_table, out_path, error_messages):
    """
    Логика, которая была в MemoqTab.export_to_excel.
    Экспорт таблицы MemoQ Stats + лог ошибок.
    """
    if memoq_table.rowCount() == 0:
        return

    df_memoq = table_to_df(memoq_table)
    for col in df_memoq.columns[1:]:
        df_memoq[col] = pd.to_numeric(df_memoq[col], errors='coerce').fillna(0).astype(int)

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df_memoq.to_excel(writer, sheet_name='MemoQ Stats', index=False)

        ws = writer.sheets['MemoQ Stats']
        row_count = df_memoq.shape[0]
        col_count = df_memoq.shape[1]

        # Заголовки (жирные, центр)
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

        # Нумерика со 2-го столбца
        for i in range(2, col_count + 1):
            for row_idx in range(2, row_count + 2):
                cell = ws.cell(row=row_idx, column=i)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')

        # Автоширина
        for col in range(1, col_count + 1):
            col_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, row_count + 2):
                val = ws.cell(row=row, column=col).value
                if val is None:
                    continue
                length = len(str(val))
                if length > max_length:
                    max_length = length
            ws.column_dimensions[col_letter].width = max_length + 2

    # Пишем лог ошибок
    if error_messages:
        base, ext = os.path.splitext(out_path)
        log_file = base + "_errors.log"
        try:
            with open(log_file, 'w', encoding='utf-8') as lf:
                for msg in error_messages:
                    lf.write(msg + "\n")
        except Exception as e:
            print(f"Не удалось записать лог-файл: {e}")